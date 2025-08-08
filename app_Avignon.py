import streamlit as st
import pandas as pd
import datetime
import io
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup
from collections import deque
import pandas.api.types as ptypes
import gspread
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from google.oauth2.service_account import Credentials
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, GridUpdateMode
from io import BytesIO
import uuid
import math
import hashlib
import json
import numpy as np
import time
from streamlit_javascript import st_javascript
import unicodedata
from urllib.parse import quote_plus

# Debug
DEBUG_TRACE_MODE = False
DEBUG_TRACE_TYPE = ["all"]
def debug_trace(trace, trace_type=["all"]):
    trace_type_requested = [s.lower() for s in DEBUG_TRACE_TYPE]
    trace_type = [s.lower() for s in trace_type]
    if DEBUG_TRACE_MODE and ("all" in trace_type_requested or any(x in trace_type_requested for x in trace_type)):
        print(trace) 

# Variables globales
BASE_DATE = datetime.date(2000, 1, 1)
MARGE = datetime.timedelta(minutes=30)
PAUSE_DEJ_DEBUT_MIN = datetime.time(11, 0)
PAUSE_DEJ_DEBUT_MAX = datetime.time(14, 0)
PAUSE_DIN_DEBUT_MIN = datetime.time(19, 0)
PAUSE_DIN_DEBUT_MAX = datetime.time(21, 0)
DUREE_REPAS = datetime.timedelta(hours=1)
DUREE_CAFE = datetime.timedelta(minutes=30)
MAX_HISTORIQUE = 20

COLONNES_ATTENDUES = ["Date", "Debut", "Fin", "Duree", "Activite", "Lieu", "Relache", "Reserve", "Priorite", "Commentaire"]
COLONNES_ATTENDUES_ACCENTUEES = ["Date", "Début", "Fin", "Durée", "Activité", "Lieu", "Relâche", "Réservé", "Priorité", "Commentaire"]
COLONNES_ATTENDUES_CARNET_ADRESSES = ["Nom", "Adresse"]

RENOMMAGE_COLONNES = {
    "Debut": "Début",
    "Duree": "Durée",
    "Reserve": "Réservé",
    "Priorite": "Prio",
    "Relache": "Relâche",
    "Activite": "Activité",
}

RENOMMAGE_COLONNES_INVERSE = {
    "Début": "Debut",
    "Durée": "Duree",
    "Réservé": "Reserve",
    "Prio": "Priorite",
    "Relâche": "Relache",
    "Activité": "Activite",
}

# Palette de couleurs
PALETTE_COULEURS_JOURS = {
    1: "#fce5cd",   2: "#fff2cc",   3: "#d9ead3",   4: "#cfe2f3",   5: "#ead1dc",
    6: "#f4cccc",   7: "#fff2cc",   8: "#d0e0e3",   9: "#f9cb9c",  10: "#d9d2e9",
    11: "#c9daf8",  12: "#d0e0e3",  13: "#f6b26b",  14: "#ffe599",  15: "#b6d7a8",
    16: "#a2c4c9",  17: "#b4a7d6",  18: "#a4c2f4",  19: "#d5a6bd",  20: "#e6b8af",
    21: "#fce5cd",  22: "#fff2cc",  23: "#d9ead3",  24: "#cfe2f3",  25: "#ead1dc",
    26: "#f4cccc",  27: "#d9d2e9",  28: "#b6d7a8",  29: "#d5a6bd",  30: "#f6b26b",
    31: "#d0e0e3"
}

# AgGrid uodate_on
AGGRID_UPDATE_ON_MENU_ACTIVITE_DOUBLE = ["modelChanged", "selectionChanged"]
AGGRID_UPDATE_ON_MENU_ACTIVITE_UNIQUE = ["modelChanged", "selectionChanged"] #, "headerFocused"] 
MENU_ACTIVITE_UNIQUE = True

LABEL_BOUTON_NOUVEAU = "Nouveau"
LABEL_BOUTON_SAUVEGARDER = "Sauvegarder"
LABEL_BOUTON_DEFAIRE = "Défaire"
LABEL_BOUTON_REFAIRE = "Refaire"
LABEL_BOUTON_AJOUTER = "Nouvelle activité"
LABEL_BOUTON_SUPPRIMER = "Supprimer"
LABEL_BOUTON_CHERCHER_WEB = "Web"
LABEL_BOUTON_CHERCHER_ITINERAIRE = "Itinéraire"
LABEL_BOUTON_PROGRAMMER = "Programmer"
LABEL_BOUTON_REPROGRAMMER = "Reprogrammer"
LABEL_BOUTON_DEPROGRAMMER = "Déprogrammer"
LABEL_BOUTON_VALIDER = "Valider"
LABEL_BOUTON_ANNULER = "Annuler"
LABEL_BOUTON_EDITER = "Editer"

CENTRER_BOUTONS = True

# 🎨 Palette de styles pour les boutons colorés
PALETTE_COULEUR_PRIMARY_BUTTONS = {
    "info":    {"bg": "#dbeafe", "text": "#0b1220"},
    "error":   {"bg": "#fee2e2", "text": "#0b1220"},
    "warning": {"bg": "#fef3c7", "text": "#0b1220"},
    "success": {"bg": "#dcfce7", "text": "#0b1220"},
}

######################
# User Sheet Manager #
######################

def get_user_id():
    params = st.query_params
    user_id_from_url = params.get("user_id", [None])

    if user_id_from_url[0]:
        st.session_state["user_id"] = user_id_from_url

    if "user_id" not in st.session_state:

        afficher_titre("Bienvenue sur le planificateur Avignon Off 👋")

        st.write("Pour commencer, clique ci-dessous pour ouvrir ton espace personnel.")
        if "new_user_id" not in st.session_state:     
            st.session_state["new_user_id"] = str(uuid.uuid4())[:8]
        new_user_id = st.session_state.new_user_id
        if st.button("Créer ma session privée"):
            st.session_state["user_id"] = new_user_id
            st.query_params.update(user_id=new_user_id)
            st.rerun()  # Recharge la page avec le nouveau paramètre
        show_user_link(new_user_id)
        st.stop()

    return st.session_state["user_id"]

def show_user_link(user_id):
    app_url = "https://planifavignon-05-hymtc4ahn5ap3e7pfetzvm.streamlit.app/"  
    user_link = f"{app_url}/?user_id={user_id}"
    st.success("Voici ton lien personnel pour revenir plus tard :")
    st.code(user_link, language="text")
    st.download_button(
        label="💾 Télécharger mon lien",
        data=user_link,
        file_name=f"lien_{user_id}.txt"
    )
    
def get_gsheet_client():
    try:
        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        return gspread.authorize(creds)
    except Exception as e:
            st.error(f"Erreur de connexion à Google Sheets : {e}")
            return None
    
def get_or_create_user_gsheets(user_id, spreadsheet_id):
    gsheets = None
    client = get_gsheet_client()
    if client is not None:    
        try:
            sh = client.open_by_key(spreadsheet_id)
        except Exception as e:
            st.error(f"Impossible d'ouvrir la Google Sheet : {e}")
            st.stop()    

        sheet_names = [f"data_{user_id}", f"links_{user_id}", f"meta_{user_id}", f"adrs_{user_id}"] # Utilisation nominale en mode multiuser avec hébergement streamlit share
        # sheet_names = [f"data", f"links", f"meta", f"adrs"] # pour debugger en local 

        gsheets = {}

        for name in sheet_names:
            try:
                ws = sh.worksheet(name)
            except gspread.WorksheetNotFound:
                ws = sh.add_worksheet(title=name, rows=1000, cols=20)
            gsheets[name.split("_")[0]] = ws  # 'data', 'links', 'meta', 'adrs'

    return gsheets

####################
# API Google Sheet #
####################

# 📥 Charge les infos persistées depuis la Google Sheet
def charger_contexte_depuis_gsheet():
    if "gsheets" not in st.session_state:
        try:
            user_id = get_user_id()
            gsheets = get_or_create_user_gsheets(user_id, spreadsheet_id="1ytYrefEPzdJGy5w36ZAjW_QQTlvfZ17AH69JkiHQzZY")
            st.session_state.gsheets = gsheets
            worksheet = gsheets["data"]
            df = get_as_dataframe(worksheet, evaluate_formulas=True)
            df.dropna(how="all")
            if len(df) > 0:

                worksheet = gsheets["links"]
                rows = worksheet.get_all_values()
                lnk = {}
                if len(rows) > 1:
                    data_rows = rows[1:]
                    lnk = {row[0]: row[1] for row in data_rows if len(row) >= 2}

                worksheet = gsheets["meta"]
                fn  = worksheet.acell("A1").value
                fp  = worksheet.acell("A2").value
                if fp is None or str(fp).strip() == "":
                    wb = download_excel_from_dropbox(fp)

                worksheet = gsheets["adrs"]
                ca = get_as_dataframe(worksheet, evaluate_formulas=True)

                df = nettoyer_donnees(df)
                initialiser_etat_contexte(df, wb, fn, lnk, ca)
                undo_redo_init(verify=False)
            else:
                initialiser_nouveau_contexte(avec_sauvegarde=False)
                undo_redo_init(verify=False)
        except Exception as e:
            pass

# 📤 Sauvegarde le DataFrame dans la Google Sheet
def sauvegarder_df_ds_gsheet(df: pd.DataFrame):
    if "gsheets" in st.session_state and st.session_state.gsheets is not None:
        try:
            gsheets = st.session_state.gsheets
            worksheet = gsheets["data"]
            worksheet.clear()
            set_with_dataframe(worksheet, df)
        except Exception as e:
            pass

# 📤 Sauvegarde les hyperliens dans la Google Sheet
def sauvegarder_lnk_ds_gsheet(lnk):
    if "gsheets" in st.session_state and st.session_state.gsheets is not None:
        try:
            gsheets = st.session_state.gsheets
            worksheet = gsheets["links"]
            worksheet.clear()
            rows = [[k, v] for k, v in lnk.items()]
            worksheet.update(range_name="A1", values=[["Clé", "Valeur"]] + rows)
        except Exception as e:
            pass

# Sauvegarde une ligne dans la Google Sheet
def sauvegarder_row_ds_gsheet(df, index_df):
    
    def convert_cell_value(x):
        if pd.isna(x):
            return ""
        elif isinstance(x, (pd.Timedelta, datetime.timedelta)):
            # Convertir en durée lisible : "1:30:00" ou minutes
            return str(x)
        elif isinstance(x, (pd.Timestamp, datetime.datetime)):
            return x.strftime("%Y-%m-%d %H:%M:%S")
        elif hasattr(x, "item"):
            return x.item()  # Pour np.int64, np.float64, etc.
        else:
            return x
    
    if "gsheets" in st.session_state and st.session_state.gsheets is not None:
        try:
            gsheets = st.session_state.gsheets
            worksheet = gsheets["data"]
            valeurs = df.drop(columns=["Debut_dt", "Duree_dt"]).loc[index_df].map(convert_cell_value).tolist()
            ligne_sheet = index_df + 2  # +1 pour index, +1 pour en-tête
            worksheet.update(range_name=f"A{ligne_sheet}", values=[valeurs])
        except Exception as e:
            pass

# 📤 Sauvegarde l'ensemble des infos persistées dans la Google Sheet
def sauvegarder_contexte_ds_gsheet(df: pd.DataFrame, lnk, fd=None, ca=None):
    if "gsheets" in st.session_state and st.session_state.gsheets is not None:
        try:
            gsheets = st.session_state.gsheets

            worksheet = gsheets["data"]
            worksheet.clear()
            set_with_dataframe(worksheet, df)

            worksheet = gsheets["links"]
            worksheet.clear()
            rows = [[k, v] for k, v in lnk.items()]
            worksheet.update(range_name="A1", values=[["Clé", "Valeur"]] + rows)

            worksheet = gsheets["meta"]
            if fd is not None:
                worksheet.update_acell("A1", fd.name)
                fp = upload_excel_to_dropbox(fd.getvalue(), fd.name)
                worksheet.update_acell("A2", fp)
                return fp
            else:
                worksheet.update_acell("A1", "")
                worksheet.update_acell("A2", "")

            worksheet = gsheets["adrs"]
            worksheet.clear()
            set_with_dataframe(worksheet, ca)

        except Exception as e:
            pass
    else:
        return ""

####################
# API Google Drive #
####################

# Sauvegarde sur le google drive le fichier Excel de l'utilisateur (nécessite un drive partagé payant sur Google Space -> Non utilisé, remplacé par DropBox)
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
# from googleapiclient.http import MediaIoBaseUpload
# from googleapiclient.discovery import build

# Id du drive partagé utilisé pour enregistrer une copie du fichier Excel utilisateur (nécessite un drive partagé payant sur Google Space -> Non utilisé, remplacé par DropBox)
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
# SHARED_DRIVE_ID = "xxxx" 

# def upload_excel_to_drive(file_bytes, filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
#     try:
#         creds = get_gcp_credentials()
#         drive_service = build("drive", "v3", credentials=creds)

#         file_metadata = {"name": filename, "parents": [SHARED_DRIVE_ID]}
#         media = MediaIoBaseUpload(BytesIO(file_bytes), mimetype=mime_type, resumable=True)
#         uploaded = drive_service.files().create(body=file_metadata, media_body=media, fields="id", supportsAllDrives=True).execute()
#         return uploaded["id"]
#     except Exception as e:
#         return None

# from googleapiclient.http import MediaIoBaseDownload

# Renvoie le fichier Excel de l'utilisateur sauvegardé sur le google drive (nécessite un drive partagé payant sur Google Space -> Non utilisé, remplacé par DropBox)
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
# def download_excel_from_drive(file_id):
#     try:
#         creds = get_gcp_credentials()
#         service = build('drive', 'v3', credentials=creds)
#         request = service.files().get_media(fileId=file_id)
#         fh = BytesIO()
#         downloader = MediaIoBaseDownload(fh, request)
#         done = False
#         while not done:
#             _, done = downloader.next_chunk()
#         fh.seek(0)
#         return load_workbook(BytesIO(fh.read()))
#     except Exception as e:
#         return Workbook()

###############
# API DropBox #
###############

import dropbox
from io import BytesIO

# Retourne les credentials pour les API Dropbox
def get_dropbox_client():
    access_token = st.secrets["dropbox"]["access_token"]
    return dropbox.Dropbox(access_token)

# Sauvegarde sur Dropbox le fichier Excel de l'utilisateur 
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def upload_excel_to_dropbox(file_bytes, filename, dropbox_path="/uploads/"):
    dbx = get_dropbox_client()
    full_path = f"{dropbox_path}{filename}"

    try:
        dbx.files_upload(file_bytes, full_path, mode=dropbox.files.WriteMode("overwrite"))
        # st.success(f"✅ Fichier '{filename}' uploadé dans Dropbox à {full_path}")
        return full_path
    except Exception as e:
        # st.error(f"❌ Erreur d’upload : {e}")
        return ""

# Renvoie le fichier Excel de l'utilisateur sauvegardé sur DropBox
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def download_excel_from_dropbox(file_path):
    dbx = get_dropbox_client()
    try:
        metadata, res = dbx.files_download(file_path)
        file_bytes = BytesIO(res.content)
        return load_workbook(file_bytes)
    except Exception as e:
        # st.error(f"❌ Erreur lors du téléchargement depuis Dropbox : {e}")
        return Workbook()

#######################
# Gestion Undo / Redo #
#######################

# Initialise les listes d'undo redo
def undo_redo_init(verify=True):
    if "historique_undo" not in st.session_state or "historique_redo" not in st.session_state or not verify:
        st.session_state.historique_undo = deque(maxlen=MAX_HISTORIQUE)
        st.session_state.historique_redo = deque(maxlen=MAX_HISTORIQUE)

# Sauvegarde du contexte courant
def undo_redo_save():
    snapshot = {
        "df": st.session_state.df.copy(deep=True),
        "liens": st.session_state.liens_activites.copy(),
        "activites_programmees_selected_row": st.session_state.activites_programmees_selected_row,
        "activites_non_programmees_selected_row": st.session_state.activites_non_programmees_selected_row,
        "menu_activites": st.session_state.menu_activites
    }
    st.session_state.historique_undo.append(snapshot)
    st.session_state.historique_redo.clear()

# Undo
def undo_redo_undo():
    if st.session_state.historique_undo:
        current = {
            "df": st.session_state.df.copy(deep=True),
            "liens": st.session_state.liens_activites.copy(),
            "activites_programmees_selected_row": st.session_state.activites_programmees_selected_row,
            "activites_non_programmees_selected_row": st.session_state.activites_non_programmees_selected_row,
            "menu_activites": st.session_state.menu_activites
        }
        st.session_state.historique_redo.append(current)
        
        snapshot = st.session_state.historique_undo.pop()
        st.session_state.df = snapshot["df"]
        st.session_state.liens_activites = snapshot["liens"]
        st.session_state.activites_programmees_selected_row = snapshot["activites_programmees_selected_row"]
        st.session_state.activites_non_programmees_selected_row = snapshot["activites_non_programmees_selected_row"]
        st.session_state.menu_activites = snapshot["menu_activites"]
        forcer_reaffichage_activites_programmees()
        forcer_reaffichage_activites_non_programmees()
        forcer_reaffichage_df("creneaux_disponibles")
        sauvegarder_df_ds_gsheet(st.session_state.df)
        st.rerun()

# Redo
def undo_redo_redo():
    if st.session_state.historique_redo:
        current = {
            "df": st.session_state.df.copy(deep=True),
            "liens": st.session_state.liens_activites.copy(),
            "activites_programmees_selected_row": st.session_state.activites_programmees_selected_row,
            "activites_non_programmees_selected_row": st.session_state.activites_non_programmees_selected_row
        }
        st.session_state.historique_undo.append(current)
        
        snapshot = st.session_state.historique_redo.pop()
        st.session_state.df = snapshot["df"]
        st.session_state.liens_activites = snapshot["liens"]
        st.session_state.activites_programmees_selected_row = snapshot["activites_programmees_selected_row"]
        st.session_state.activites_non_programmees_selected_row = snapshot["activites_non_programmees_selected_row"]
        st.session_state.menu_activites = snapshot["menu_activites"]
        forcer_reaffichage_activites_programmees()
        forcer_reaffichage_activites_non_programmees()
        forcer_reaffichage_df("creneaux_disponibles")
        sauvegarder_df_ds_gsheet(st.session_state.df)
        st.rerun()

#########################
# Fonctions utilitaires #
#########################

# Essai de boutons html (à creuser -> permettrait d'avoir des boutons horizontaux avec grisés sur mobile)
def boutons_html():
    # Images
    undo_icon = image_to_base64("undo_actif.png")
    undo_disabled_icon = image_to_base64("undo_inactif.png")
    redo_icon = image_to_base64("undo_actif.png")
    redo_disabled_icon = image_to_base64("undo_inactif.png")

    # États
    undo_enabled = True
    redo_enabled = False

    # Lire le paramètre ?btn=undo ou ?btn=redo
    params = st.query_params
    clicked_btn = params.get("btn", None)

    # Action déclenchée
    if clicked_btn == "undo":
        st.success("Undo cliqué ✅")
        undo_redo_undo()

    elif clicked_btn == "redo":
        st.success("Redo cliqué ✅")
        undo_redo_redo()

    # Affichage des boutons côte à côte (même taille, même style)
    html = f"""
    <div style="display: flex; gap: 1em; align-items: center;">
    <a href="?btn=undo">
        <button style="background:none;border:none;padding:0;cursor:{'pointer' if undo_enabled else 'default'};" {'disabled' if not undo_enabled else ''}>
        <img src="data:image/png;base64,{undo_icon if undo_enabled else undo_disabled_icon}" width="32">
        </button>
    </a>
    <a href="?btn=redo">
        <button style="background:none;border:none;padding:0;cursor:{'pointer' if redo_enabled else 'default'};" {'disabled' if not redo_enabled else ''}>
        <img src="data:image/png;base64,{redo_icon if redo_enabled else redo_disabled_icon}" width="32">
        </button>
    </a>
    </div>
    """

    st.markdown(html, unsafe_allow_html=True)

# Evite la surbrillance rose pâle des lignes qui ont le focus sans être sélectionnées dans les AgGrid
def patch_aggrid_css():
    # st.markdown("""
    #     <style>
    #     .ag-row.ag-row-focus {
    #         background-color: transparent !important;
    #     }
    #     </style>
    # """, unsafe_allow_html=True)
    st.markdown("""
        <style>
        /* Supprime l'effet de hover pâle sur mobile */
        .ag-row:hover:not(.ag-row-selected) {
            background-color: transparent !important;
        }

        /* Supprime l'effet de "focus ligne" qui donne le rose pâle */
        .ag-row.ag-row-focus:not(.ag-row-selected) {
            background-color: transparent !important;
        }

        /* Ne touche pas aux lignes sélectionnées (rose plus foncé) */
        </style>
    """, unsafe_allow_html=True)

# Renvoie True si l'appli tourne sur mobile (ne fonctionne pas...)
def mode_mobile():    
    from streamlit_js_eval import streamlit_js_eval, get_geolocation
    if "mode_mobile" not in st.session_state:
        _mode_mobile = False
        user_agent = streamlit_js_eval(js_expressions="navigator.userAgent", key="ua")
        if user_agent: # Renvoie toujours None...
            if "Mobile" in user_agent or "Android" in user_agent or "iPhone" in user_agent:
                _mode_mobile = True
        st.session_state.mode_mobile = _mode_mobile
    return True # st.session_state.mode_mobile

import streamlit as st

# Injecte le CSS permettent de colorer les primary buttons selon les styles de PALETTE_COULEUR_PRIMARY_BUTTONS ("info", "error", etc.) 
def injecter_css_pour_primary_buttons(type_css):
    palette = PALETTE_COULEUR_PRIMARY_BUTTONS.get(type_css, PALETTE_COULEUR_PRIMARY_BUTTONS["info"])
    st.markdown(f"""
    <style>
    button[data-testid="stBaseButton-primary"]{{
    background-color: {palette["bg"]} !important;   /* fond info */
    color: #0b1220 !important;
    border: none !important;                /* supprime toutes les bordures */
    outline: none !important;
    box-shadow: none !important;
    text-align: left !important;
    width: 100% !important;
    padding: 0.9em 1em !important;
    border-radius: 0.5em !important;
    white-space: normal !important;
    line-height: 1.4 !important;
    cursor: pointer !important;
    }}
    </style>
    """, unsafe_allow_html=True)

# Affiche l'équivalent d'un st.info ou st.error avec un label
# Si key est fourni, un bouton clickable de type primary est utilisé 
# Ce bouton doit être stylé avec un CSS ciblant les boutons de type primary et injecté par l'appelant pour être coloré correctement
# Mais attention tous les boutons de type primary seront alors stylés de la même manière 
def st_info_error_avec_label(label, info_text, key=None, color="blue", afficher_label=True, label_separe=True):
    
    def st_info_error_ou_bouton(label, info_text, key, color):
        if key:
            return st.button(info_text, key=key, type="primary", use_container_width=True)
        else:
            if color.lower() == "red":
                st.error(info_text) 
            else:
                st.info(info_text
                               )
    if label_separe:
        if afficher_label:
            st.markdown(f"""
            <div style='
                font-size: 0.88rem;
                font-weight: normal;
                margin-bottom: 0.2rem;
            '>
                {label}
            </div>
            """, unsafe_allow_html=True)

        return st_info_error_ou_bouton(label, info_text, key, color)
    else:
        info_text = f"**{label}:** {info_text}" if afficher_label else info_text
        return st_info_error_ou_bouton(label, info_text, key, color)

# Cast en int sûr
def safe_int(val, default=None):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

# Indique si val est un float valide
def est_float_valide(val):
    return (isinstance(val, float) or isinstance(val, int) or  isinstance(val, np.float64) or isinstance(val, np.int64)) and not math.isnan(val)
    
# Renvoie val sous la forme "10h00" si datetime ou time, "" si None, str(val).strip() sinon
def heure_str(val):
    from datetime import datetime, time
    if isinstance(val, (datetime, time)):
        return val.strftime("%Hh%M")
    if pd.isna(val):
        return ""
    return str(val).strip()

# Renvoie un datetime basé sur BASE_DATE si h est datetime, time, str de la forme 10h00, 10:00 ou 10:00:00, None dans les autres cas
def heure_parse(h):
    from datetime import datetime, time

    if pd.isna(h) or str(h).strip() == "":
        return datetime.combine(BASE_DATE, time(0, 0))  # Heure nulle par défaut        if isinstance(h, time):
    
    if isinstance(h, datetime):
        return datetime.combine(BASE_DATE, h.time())
    
    h_str = str(h).strip()

    # Format 10h00
    if re.match(r"^\d{1,2}h\d{2}$", h_str):
        try:
            return datetime.strptime(f"{BASE_DATE.isoformat()} {h_str}", "%Y-%m-%d %Hh%M")
        except ValueError:
            return None

    # Format 10:00 ou 10:00:00
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", h_str):
        try:
            t = datetime.strptime(h_str, "%H:%M").time()
            return datetime.combine(BASE_DATE, t)
        except ValueError:
            try:
                t = datetime.strptime(h_str, "%H:%M:%S").time()
                return datetime.combine(BASE_DATE, t)
            except ValueError:
                return None

    return None

# Indique si une valeur à un format heure semblable à 10h00
def est_heure_valide(val):
    if pd.isna(val):
        return False
    try:
        return re.fullmatch(r"\d{1,2}h\d{2}", val.strip()) if val else False
    except Exception:
        return False
    
# Renvoie val sous la forme "1h00" si timedelta, "" si None, str(val).strip() sinon
def duree_str(val):
    from datetime import datetime, time
    if pd.isna(val):
        return ""
    if isinstance(val, pd.Timedelta):
        total_minutes = int(val.total_seconds() // 60)
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h}h{m:02d}"
    return str(val).strip()

# Renvoie un timedelta si h est timedelta, datetime, time, str de la forme 1h00, 1:00 ou 1:00:00, None dans les autres cas
def duree_parse(d):
    from datetime import datetime, time

    if pd.isna(d) or str(d).strip() == "":
        return pd.Timedelta(0)

    # Si c'est déjà un timedelta
    if isinstance(d, pd.Timedelta):
        return d

    # Si c'est un datetime.time
    if isinstance(d, time):
        return pd.Timedelta(hours=d.hour, minutes=d.minute, seconds=d.second)

    # Si c'est un datetime.datetime
    if isinstance(d, datetime):
        t = d.time()
        return pd.Timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)

    d_str = str(d).strip().lower()

    # Format "1h30"
    if re.match(r"^\d{1,2}h\d{2}$", d_str):
        h, m = map(int, d_str.replace("h", " ").split())
        return pd.Timedelta(hours=h, minutes=m)

    # Format "1:30" ou "1:30:00"
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", d_str):
        try:
            parts = list(map(int, d_str.split(":")))
            if len(parts) == 2:
                h, m = parts
                return pd.Timedelta(hours=h, minutes=m)
            elif len(parts) == 3:
                h, m, s = parts
                return pd.Timedelta(hours=h, minutes=m, seconds=s)
        except ValueError:
            return None

    return None

# Indique si une valeur à un format durée semblable à 1h00
def est_duree_valide(val):
    if pd.isna(val):
        return False
    try:
        return re.fullmatch(r"\d{1,2}h[0-5]\d", val.strip()) is not None if val else False
    except Exception:
        return False
    

# Calcule l'heure de fin à partir de l'heure de début et de la durée    
def calculer_fin(h, d, fin_actuelle=""):
    if isinstance(d, pd.Timedelta) and not pd.isna(h):
        total = h + d
        return f"{total.hour:02d}h{total.minute:02d}"
    else:
        return fin_actuelle if pd.notna(fin_actuelle) else ""

# Calcule l'heure de fin à partir d'une row
def calculer_fin_row(row):
    h = row.get("Debut_dt")
    d = row.get("Duree_dt")
    fin_actuelle = row.get("Fin")
    return calculer_fin(h, d, fin_actuelle)

# Formatte un objet timedelta en une chaîne de caractères "XhYY"
def formatter_timedelta(d):
    if isinstance(d, datetime.timedelta):
        total_seconds = int(d.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours}h{minutes:02d}"
    return d
    
# Formatte le contenu d'une cellule entiere
def formatter_cellule_int(d):
    if isinstance(d, int) or isinstance(d, float):
        if isinstance(d, float) and math.isnan(d):
            return d
        return int(d)
    return d
    
# Renvoie une bitmap encodée en format Base64 à partir d'un fichier
import base64
def image_to_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

# Ajoute les colonnes non présentes du df dans une row (hors colonnes de travail ["Debut_dt", "Duree_dt"])
def completer_ligne(ligne_partielle):
    colonnes_df_utiles = [col for col in st.session_state.df if col not in ["Debut_dt", "Duree_dt"]]
    colonnes_supplementaires = [col for col in ligne_partielle.keys() if col not in colonnes_df_utiles]
    colonnes_finales = colonnes_df_utiles + colonnes_supplementaires
    return {col: ligne_partielle.get(col, None) for col in colonnes_finales}

# Selectbox avec items non editables (contrairement à st.selectbox())
def selectbox_aggrid(label, options, key="aggrid_selectbox", height=100):
    df = pd.DataFrame({"Choix": [options[0]]})
    
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_column(
        "Choix",
        editable=True,
        cellEditor="agSelectCellEditor",
        cellEditorParams={"values": options},
        singleClickEdit=True,
        minWidth=120  # 🔧 largeur minimale lisible
    )
    gb.configure_grid_options(domLayout='autoHeight')
    gb.configure_grid_options(onGridReady="""
        function(params) {
            setTimeout(function() {
                params.api.sizeColumnsToFit();
            }, 100);
        }
    """)
    gridOptions = gb.build()

    st.markdown(f"{label}")
    response = AgGrid(
        df,
        gridOptions=gridOptions,
        height=height,
        key=key,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=True
    )

    try:
        return response["data"]["Choix"].iloc[0]  # ✅ corrige le warning
    except:
        return None  # En cas de suppression accidentelle

# Force le reaffichage d'un dataframe
def forcer_reaffichage_df(key):
    session_state_reset_counter = key + "_reset_counter"
    if session_state_reset_counter in st.session_state:
        st.session_state[session_state_reset_counter] += 1 
    session_state_forcer_reaffichage = key + "_forcer_reaffichage"
    if session_state_forcer_reaffichage in st.session_state:
        st.session_state[session_state_forcer_reaffichage] = True

# Affichage d'un dataframe
def afficher_df(label, df, hide=[], key="affichage_df", colorisation=False):

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 150
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de savoir si l'on doit forcer le réaffichage de l'aggrid après une suppression de ligne 
    session_state_reset_counter = key + "_reset_counter"
    if session_state_reset_counter not in st.session_state:
        st.session_state[session_state_reset_counter] = 0
    
    # Initialisation du flag permettant de savoir si l'on est en mode réaffichage complet de l'aggrid
    session_state_forcer_reaffichage = key + "_forcer_reaffichage"
    if session_state_forcer_reaffichage not in st.session_state:
        st.session_state[session_state_forcer_reaffichage] = False
   
    
    gb = GridOptionsBuilder.from_dataframe(df)

    #Colonnes cachées
    for col in hide:
        if col in df.columns:
            gb.configure_column(col, hide=True)

    # Colorisation
    if colorisation:
        if "Date" in df.columns:
            df["__jour"] = df["Date"].apply(lambda x: int(str(int(float(x)))[-2:]) if pd.notna(x) else None)
            gb.configure_column("__jour", hide=True)
            gb.configure_grid_options(getRowStyle=JsCode(f"""
            function(params) {{
                const jour = params.data.__jour;
                const couleurs = {PALETTE_COULEURS_JOURS};
                if (jour && couleurs[jour]) {{
                    return {{ 'backgroundColor': couleurs[jour] }};
                }}
                return null;
            }}
            """))

    # Configuration de la sélection
    pre_selected_row = 0  # par défaut
    session_state_selected_row = key + "_selected_row"
    if session_state_selected_row in st.session_state and st.session_state[session_state_selected_row] is not None:
        selected_row_courante = st.session_state[session_state_selected_row]
        position = trouver_position_ligne(df, selected_row_courante.to_dict())
        pre_selected_row = position if position is not None else pre_selected_row
    gb.configure_selection(selection_mode="single", use_checkbox=False, pre_selected_rows=[pre_selected_row])

    # Retaillage auto des largeurs de colonnes
    gb.configure_grid_options(onGridReady=JsCode(f"""
        function(params) {{
            params.api.sizeColumnsToFit();
            params.api.ensureIndexVisible({pre_selected_row}, 'middle');
            params.api.getDisplayedRowAtIndex({pre_selected_row}).setSelected(true);
        }}
    """))

    grid_options = gb.build()
    grid_options["suppressMovableColumns"] = True

    st.markdown(f"{label}")
    response = AgGrid(
        df,
        gridOptions=grid_options,
        height=height,
        key=f"{key} {st.session_state[session_state_reset_counter]}",
        update_mode=GridUpdateMode.MODEL_CHANGED | GridUpdateMode.SELECTION_CHANGED,
        allow_unsafe_jscode=True,
        # fit_columns_on_grid_load=False
    )

    selected_rows = response["selected_rows"]
    if st.session_state[session_state_forcer_reaffichage] == True:
        row = df.iloc[pre_selected_row]
    else:
        if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
            row = selected_rows.iloc[0] 
        elif isinstance(selected_rows, list) and len(selected_rows) > 0:
            row = selected_rows[0]
        else: 
            row = df.iloc[pre_selected_row]
    st.session_state[session_state_forcer_reaffichage] = False

    st.session_state[session_state_selected_row] = row

    return row

# Renvoie le numero de ligne d'un df qui matche des valeurs
def trouver_position_ligne(df, valeurs):
    for i, row in df.iterrows():
        match = True
        for col, val in valeurs.items():
            if col in row and not pd.isna(row[col]):
                if row[col] != val:
                    match = False
                    break
        if match:
            return df.index.get_loc(i)
    return None

# Renvoie l'index de la ligne la plus proche dans un df_display d'aggrid
# Le df_display est supposé contenir dans la colonne __index l'index du df de base
def ligne_voisine_index(df_display, index_df):
    df_display_reset = df_display.reset_index(drop=True)
    if pd.notna(index_df):
        if len(df_display_reset) > 0:
            selected_row_pos = df_display_reset["__index"].eq(index_df).idxmax() 
            new_selected_row_pos = selected_row_pos + 1 if  selected_row_pos + 1 <= len(df_display) - 1 else max(selected_row_pos - 1, 0)
            return df_display_reset.iloc[new_selected_row_pos]["__index"] 
    else:
        return None

# Selectbox avec items non editables (contrairement à st.selectbox())
def aggrid_single_selection_list(label, choices, key="aggrid_select", hauteur=200):
    # Garde-fou : le label doit être une chaîne
    if not isinstance(label, str):
        raise ValueError(f"Le paramètre `label` doit être une chaîne, reçu : {type(label)}")

    # Transformation si liste de listes
    if choices and isinstance(choices[0], (list, tuple)):
        choices = [" | ".join(map(str, ligne)) for ligne in choices]

    df = pd.DataFrame({label: choices})

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_selection(selection_mode="single", use_checkbox=False)
    gb.configure_grid_options(
        domLayout='normal',
        headerHeight=0,
        suppressRowHoverHighlight=True,
        suppressCellFocus=True,
    )
    gb.configure_column(label, header_name="", wrapText=True, autoHeight=True, minWidth=200, flex=1)

    st.markdown("""
        <style>
        .ag-root-wrapper,
        .ag-theme-streamlit,
        .ag-header,
        .ag-cell:focus,
        .ag-cell,
        .ag-row-hover {
            border: none !important;
            outline: none !important;
            box-shadow: none !important;
            background-color: transparent !important;
        }
        .ag-header { display: none !important; }
        </style>
    """, unsafe_allow_html=True)

    response = AgGrid(
        df,
        gridOptions=gb.build(),
        key=key,
        height=hauteur,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
        enable_enterprise_modules=False
    )

    selected = response.get("selected_rows")
    if selected:
        valeur = selected[0][label]
        index_selection = selected[0].get("_selectedRowNodeInfo", {}).get("nodeRowIndex", None)
        return valeur
    else:
        return choices[0]

# Crée un hash des colonnes d'un df et de parametres.
def hash_df(df: pd.DataFrame, colonnes_a_garder: list=None, colonnes_a_enlever: list=None, params=None):
    # Attention : convertir les colonnes de type datetime en string pour JSON
    if colonnes_a_garder is None:
        df_subset = df
    else:
        df_subset = df[colonnes_a_garder]
    
    if colonnes_a_enlever is not None:
        df_subset = df_subset.drop(colonnes_a_enlever, axis=1)

    df_subset = df_subset.astype(str)
    
    data = {
        "df": df_subset.to_dict("records"),
        "params": params
    }
    json_data = json.dumps(data, sort_keys=True)
    return hashlib.sha256(json_data.encode()).hexdigest()

##########################
# Fonctions applicatives #
##########################

# Renvoie un descripteur d'activité à partir d'une date et d'une ligne du df
def get_descripteur_activite(date, row):
    titre = f"{date} - [{row['Debut'].strip()} - {row['Fin'].strip()}] - {row['Activite']}"
    if not (pd.isna(row["Lieu"]) or str(row["Lieu"]).strip() == ""):
        titre = titre + f"( {row['Lieu']}) - P{formatter_cellule_int(row['Priorite'])}"
    return titre

# Affiche le titre de la page de l'application
def afficher_titre(title):
    # Réduire l’espace en haut de la page
    st.markdown(
        """
        <style>
            .block-container {
                padding-top: 2rem;
            }
        </style>
        """, unsafe_allow_html=True
    )

    # Titre de la page
    st.markdown(f"## {title}")

# Affiche l'aide de l'application
def afficher_aide():
    with st.expander("À propos"):
    
        with st.expander("Fonctionnalités générales"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Cette application offre les fonctionnalités suivantes:</p>
            <ul style="margin-top: 0em; margin-bottom: 2em">
            <li>Choix de la période à programmer</li>
            <li>Chargement des activités à programmer à partir d'un fichier Excel</li>
            <li>Gestion de la programmation des activités en respectant les règles décrites dans le paragraphe ci-dessous</li>
            <li>Gestion des créneaux disponibles</li>
            <li>Prise en compte optionnelle des pauses (déjeuner, dîner, café)</li>
            <li>Gestion des liens de recherche sur le net</li>
            <li>Sauvegarde des données modifiées dans un fichier téléchargeable</li>
            <li>Fonction défaire / refaire</li>
            <li>Vérification de cohérence des données d'entrée (chevauchements d'activités, marges trop courtes, formats de données)</li>
            </ul>            
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Règles de programmation des activités"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Règles adoptées pour la programmation des activités:</p>
            <ul style="margin-top: 0em; margin-bottom: 2em">
            <li>30 minutes de marge entre activités</li>
            <li>1 heure par pause repas</li>
            <li>1/2 heure par pause café sans marge avec l'activité précédente ou suivante</li>
            <li>Respect des jours de relâches</li>
            </ul>
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Utilisation"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p>Les activités à programmer sont présentées dans deux tableaux séparés, 
                l'un pour les activités déja programmées à une date donnée, l'autre pour les activités restant à programmer. 
                Dans ces deux tableaux les informations sont éditables, sauf les heures de fin (qui sont calculées automatiquement) 
                et les date de programmation, heure de début et durée des activités réservées (celles dont la colonne 'Réservé' est à Oui).</p>
            
            <p>Deux autres tableaux adressent la gestion des créneaux disponibles. 
                Le premier présente les créneaux encore disponibles sur la période considérée et le deuxième les activités programmables dans le créneau sélectionné en tenant compte de leur durée et de la marge entre activités.</p>
            
            <p style="margin-bottom: 0.2em">Les menus sont regroupés dans une barre latérale escamotable:</p>
            <ul style="margin-top: 0em">
                <li>Menu Fichier: permet de charger un contexte à partir d'un fichier, initialiser un nouveau contexte, sauvegarder le contexte courant dans un fichier téléchargeable.</li>
                <li>Menu Edition: permet de défaire, refaire une opération.</li>
                <li>Menu Activités: permet sur l'activité séléctionnée dans les tableaux d'activites programmées et non programmées (vous pouvez passer de l'activité sélectionnée dans l'un ou l'autre des tableaux en cliquant sur le champ affichant l'activité courante) de:
                        <ul>
                        <li>rechercher de l'information sur le Web (via un lien Web éditable dans les propriétés),</li> 
                        <li>rechercher un itinaire, sur la base du lieu enregistré pour l'activité (l'application d'itinéraire et la ville de recherche par défaut sont réglables dans la section Paramètres et un carnet d'adresses avec colonnes Nom et Adresse peut être enregistré dans la feuille 2 du fichier Excel d'entrée),</li>
                        <li>supprimer l'activité (si elle n'est pas réservée),</li> 
                        <li>déprogrammer l'activité (si elle est déjà programmée sans être réservée),</li>
                        <li>programmer / reprogrammer l'activité (si elle n'est pas réservée et que d'autres dates de programmation sont possibles)</li>
                        <li>éditer les propriétés l'activité.</li>
                        </ul>
                </li>
                <li>Menu créneau disponible: présente le créneau sélectionné dans la table des créneaux disponibles, l'activité sélectionnée dans la table des activités programmables sur ce créneau et permet de programmer ladite activité sur le créneau choisi.</li>
            </ul>
                        
            <p style="margin-bottom: 0.2em">En haut de la page principale une rubrique escamotable 'Infos' présente:</p>
            <ul style="margin-top: 0em">
                <li>La présente aide.</li>
                <li>Une rubrique présentant les incohérences dans le fichier chargé (notamment les chevauchements de programmation en tenant compte des marges entre activités). 
                    Cette rubrique est mise à jour au fil de l'eau.</li>
                <li>La période programmation: elle est automatiquement déduite des activités renseignées dans le fichier chargé, mais peut être modifiée en cours d'édition. Par défaut l'application recherche les dates de début et de fin du festival de l'année courante.</li>
                <li>Les paramètres de l'application comprennant:
                        <ul>
                        <li>la marge entre activités</li>
                        <li>la durée des pauses repas et café</li>
                        <li>le nom de l'application d'itinéraire (Google Maps, Apple, etc.)</li>
                        <li>la ville de recherche par défaut pour la recherche d'itinéraire</li>
                        <li>la possibilité de choisir si les menus de gestion des activités sont dans la barre latérale ou la page principale.</li>
                        </ul>
                </li>
            </ul>
                        
            <p>A la première utilisation l'application propose à l'utilisateur de créer un espace personnel dans lequel est automatiquement sauvegardé le contexte de travail (l'adresse de cet espace est : adresse de l'application/?user_id=id utilisateur).
                En cas de rupture de connexion avec le serveur, le travail en cours est ainsi automatiquement restitué à la prochaine reconnexion.</p>
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Format des données"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Le fichier Excel d'entrée doit contenir en feuille 1 les colonnes suivantes:</p>
            <ul style="margin-top: 0em; margin-bottom: 2em">
            <li>Date : Date de l'activité (entier)</li>
            <li>Début : Heure de début de l'activité (format HHhMM)</li>
            <li>Fin : Heure de fin de l'activité (format HHhMM)</li>
            <li>Durée : Durée de l'activité (format HHhMM ou HHh)</li>
            <li>Activité : Nom de l'activité (nom de spectacle, pause, visite, ...)</li>
            <li>Lieu : Lieu de l'activité</li>
            <li>Relâche : Jours de relâche pour l'activité (liste d'entiers, peut être vide)</li>
            <li>Réservé : Indique si l'activité est réservée (Oui/Non, vide interpété comme Non)</li>
            </ul>

            <p>En feuille 2 peut être fourni un carnet d'adresses des lieux d'activités utilisé pour la recherche d'itinéraire. 
            Il doit comprendre au moins une colonne Nom et une colonne Adresse.</p>

            <p>📥Un modèle Excel est disponible <a href="https://github.com/jnicoloso-91/PlanifAvignon-05/raw/main/Mod%C3%A8le%20Excel.xlsx" download>
            ici
            </a></p>
            <p>ℹ️ Si le téléchargement ne démarre pas, faites un clic droit → "Enregistrer le lien sous...".</p>
            </div>
            """, unsafe_allow_html=True)  

def get_dates_festival():
    
    # 1️⃣ Tentative de récupération des dates du festival depuis le site officiel (recherche simple)
    def fetch_off_festival_dates():
        url = "https://www.festivaloffavignon.com/"
        r = requests.get(url, timeout=5)
        soup = BeautifulSoup(r.text, "html.parser")
        # Recherche dans le texte "du 5 au 26 juillet 2025"
        text = soup.get_text()
        match = re.search(r"du\s+(\d{1,2})\s+juillet\s+au\s+(\d{1,2})\s+juillet\s+2025", text, re.IGNORECASE)
        if match:
            d1, d2 = map(int, match.groups())
            base_year = 2025
            base_month = 7
            return datetime.date(base_year, base_month, d1), datetime.date(base_year, base_month, d2)
        return None, None

    if "festival_debut" not in st.session_state or "festival_fin" not in st.session_state:
        debut, fin = fetch_off_festival_dates()
        if debut and fin:
            st.session_state.festival_debut = debut
            st.session_state.festival_fin = fin
        else:
            # Valeurs de secours (manuelles)
            st.session_state.festival_debut = datetime.date(2025, 7, 5)
            st.session_state.festival_fin = datetime.date(2025, 7, 26)
    return {
        "debut": st.session_state.festival_debut,
        "fin": st.session_state.festival_fin
    }

# Affichage de la période à programmer
def initialiser_periode_a_programmer(df):

    if "nouveau_fichier" not in st.session_state:
        st.session_state.nouveau_fichier = True
    
    # Initialisation de la periode si nouveau fichier
    if st.session_state.nouveau_fichier == True:
        # Reset du flag déclenché par callback upload
        st.session_state.nouveau_fichier = False

        # Initialisation des variables de début et de fin de période à programmer
        periode_a_programmer_debut = None
        periode_a_programmer_fin = None

        # Garde uniquement les valeurs non nulles et convertibles de la colonne Date du df
        dates_valides = df["Date"].dropna().apply(lambda x: int(float(x)) if str(x).strip() != "" else None)
        dates_valides = dates_valides.dropna().astype(int)

        if not dates_valides.empty:
            # Conversion en datetime
            base_date = datetime.date(datetime.date.today().year, 7, 1)
            dates_datetime = dates_valides.apply(lambda j: datetime.datetime.combine(base_date, datetime.datetime.min.time()) + datetime.timedelta(days=j - 1))

            if not dates_datetime.empty:
                periode_a_programmer_debut = dates_datetime.min()
                periode_a_programmer_fin = dates_datetime.max()

        if periode_a_programmer_debut is None or periode_a_programmer_fin is None:
            dates_festival = get_dates_festival()
            periode_a_programmer_debut = dates_festival["debut"]
            periode_a_programmer_fin = dates_festival["fin"]
        
        st.session_state.periode_a_programmer_debut = periode_a_programmer_debut
        st.session_state.periode_a_programmer_fin = periode_a_programmer_fin
    
    if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
        dates_festival = get_dates_festival()
        st.session_state.periode_a_programmer_debut = dates_festival["debut"]
        st.session_state.periode_a_programmer_fin = dates_festival["fin"]

# Affichage de la période à programmer
def afficher_periode_a_programmer(df):
    if "periode_a_programmer_debut" in st.session_state and "periode_a_programmer_fin" in st.session_state:
        with st.expander("Période de programmation"):
            col1, col2 = st.columns(2)
            with col1:
                st.session_state.periode_a_programmer_debut = st.date_input("Début", value=st.session_state.periode_a_programmer_debut, format="DD/MM/YYYY")
            with col2:
                st.session_state.periode_a_programmer_fin = st.date_input("Fin", value=st.session_state.periode_a_programmer_fin, format="DD/MM/YYYY")

# Affichage des paramètres
def afficher_parametres():
    with st.expander("Paramètres"):

        # Marge entre activités et durée des pauses
        if "MARGE" not in st.session_state:
            st.session_state.MARGE = MARGE

        minutes = st.slider(
            "Marge entre activités (minutes)",
            min_value=0,
            max_value=120,
            value=int(st.session_state.MARGE.total_seconds() // 60),
            step=5,
            help="Cette marge s'applique au calcul des activités programmables dans un créneau donnée. Pour les pauses café elle ne s'applique qu'à l'activité suivante ou la précédente mais pas aux deux.",
        )
        nouvelle_marge = datetime.timedelta(minutes=minutes)
        if st.session_state.MARGE != nouvelle_marge:
            st.session_state.MARGE = nouvelle_marge  
            forcer_reaffichage_df("creneaux_disponibles")
            st.rerun()   

        if "DUREE_REPAS" not in st.session_state:
            st.session_state.DUREE_REPAS = DUREE_REPAS

        minutes = st.slider(
            "Durée des pauses repas (minutes)",
            min_value=0,
            max_value=120,
            value=int(st.session_state.DUREE_REPAS.total_seconds() // 60),
            step=5,
            help="Durée appliquée à la programmation des pauses repas.",
        )
        nouvelle_duree = datetime.timedelta(minutes=minutes)
        if st.session_state.DUREE_REPAS != nouvelle_duree:
            st.session_state.DUREE_REPAS = nouvelle_duree  
            forcer_reaffichage_df("creneaux_disponibles")
            st.rerun()   

        if "DUREE_CAFE" not in st.session_state:
            st.session_state.DUREE_CAFE = DUREE_CAFE

        minutes = st.slider(
            "Durée des pauses café (minutes)",
            min_value=0,
            max_value=120,
            value=int(st.session_state.DUREE_CAFE.total_seconds() // 60),
            step=5,
            help="Durée appliquée à la programmation des pauses café.",
        )
        nouvelle_duree = datetime.timedelta(minutes=minutes)
        if st.session_state.DUREE_CAFE != nouvelle_duree:
            st.session_state.DUREE_CAFE = nouvelle_duree  
            forcer_reaffichage_df("creneaux_disponibles")
            st.rerun()   

        # Application itinéraire
        platform = get_platform()

        if platform is not None:
            # Proposer le lien adapté
            if platform == "iOS":
                options = ["Apple Maps", "Google Maps App", "Google Maps Web"]
            elif platform == "Android":
                options = ["Google Maps App", "Google Maps Web"]
            else:
                options = ["Google Maps Web"]
        
        if "itineraire_app" not in st.session_state:
            # Valeur par défaut pour l'application itinéraire
            st.session_state.itineraire_app = "Google Maps Web"
        st.selectbox("Application itinéraire", 
                    options=options, 
                    index=options.index(st.session_state.itineraire_app), 
                    key="itineraire_app", 
                    help="Sélectionnez l'application à utiliser pour la recherche d'itinéraire. Si vous utilisez un téléphone mobile, vous pouvez chosir les applications Google Maps ou Apple Maps, ou bien Google Maps Web. Si vous utilisez un ordinateur, seul Google Maps Web est proposé.")

        # Ville par défaut
        if "city_default" not in st.session_state:
            st.session_state.city_default = "Avignon"
        st.session_state.city_default = st.text_input("Ville par défaut pour la recherche d'itinéraire",
                                                      value=st.session_state.city_default,
                                                      key="city_default_input",
                                                      help="Si vide, la ville du lieu de l'activité est utilisée.")
        
        # 
        options = {"Oui": True, "Non": False}
        
        # Valeur par défaut si absente
        if "sidebar_menus" not in st.session_state:
            st.session_state.sidebar_menus = True

        choix = st.selectbox(
            "Menus activités dans barre latérale :",
            options=list(options.keys()),
            index=0 if st.session_state.sidebar_menus else 1,
            key="sidebar_menus_select"
        )

        # Mise à jour dans session_state
        st.session_state.sidebar_menus = options[choix]


# Met à jour les données calculées
def maj_donnees_calculees(df):
    try:
        df["Debut_dt"] = df["Debut"].apply(heure_parse)
        df["Duree_dt"] = df["Duree"].apply(duree_parse)
        df["Fin"] = df.apply(calculer_fin_row, axis=1)    
    except:
        pass        

# Nettoie les données du tableau Excel importé
def nettoyer_donnees(df):
    try:
        # Nettoyage noms de colonnes : suppression espaces et accents
        df.columns = df.columns.str.strip().str.replace("\u202f", " ").str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("utf-8")

        if not all(col in df.columns for col in COLONNES_ATTENDUES):
            st.error("Le fichier n'est pas au format Excel ou ne contient pas toutes les colonnes attendues: " + ", ".join(COLONNES_ATTENDUES_ACCENTUEES) + ".")
        elif (len(df) == 0):
            st.error("Le fichier est vide")
        else:

            # Suppression des lignes presque vides i.e. ne contenant que des NaN ou des ""
            df = df[~df.apply(lambda row: all(pd.isna(x) or str(x).strip() == "" for x in row), axis=1)].reset_index(drop=True)

            # Transformation de la colonne Date en entiers
            df["Date"] = pd.to_numeric(df["Date"], errors="coerce").astype("Int64")

            # Nettoyage Heure (transforme les datetime, time et None en str mais ne garantit pas le format HHhMM, voir heure_parse et est_heure_valide pour cela)
            df["Debut"] = df["Debut"].apply(heure_str)

            # Nettoyage Duree (transforme les timedelta et None en str mais ne garantit pas le format HhMM, voir duree_parse et est_duree_valide pour cela)
            df["Duree"] = df["Duree"].apply(duree_str)

            # Force le type de certaines colonnes pour éviter les erreurs de conversion ultérieures
            colonnes_cibles = {
                "Debut": "string",
                "Fin": "string",
                "Duree": "string",
                "Activite": "string",
                "Lieu": "string"
            }
            for col, dtype in colonnes_cibles.items():
                df[col] = df[col].astype(dtype) 

            df["Relache"] = df["Relache"].astype("object").fillna("").astype(str)
            df["Priorite"] = pd.to_numeric(df["Priorite"], errors="coerce").astype("Int64")

            # Valide le contexte si pas d'exception dans le traitement précédent
            if "contexte_invalide" in st.session_state:
                del st.session_state["contexte_invalide"]

        return df
            
    except Exception as e:
        st.error(f"Erreur lors du décodage du fichier : {e}")

# Renvoie les hyperliens de la colonne Activité 
def get_liens_activites(wb):
    liens_activites = {}
    try:
        ws = wb.worksheets[0]
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() in ["activité"]:
                col_excel_index = cell.column
        for row in ws.iter_rows(min_row=2, min_col=col_excel_index, max_col=col_excel_index):
            cell = row[0]
            if cell.hyperlink:
                liens_activites[cell.value] = cell.hyperlink.target
            else:
                # Construire l'URL de recherche par défaut
                if cell.value is not None:
                    url = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={cell.value.replace(' ', '+')}"
                    liens_activites[cell.value] = url  # L'enregistrer dans la session
        return liens_activites
    except:
        return liens_activites

# Vérifie la cohérence des informations du dataframe et affiche le résultat dans un expander
def verifier_coherence(df):
    
    # @st.cache_data
    def get_log_verifier_coherence(df):
        erreurs = []

        def est_entier(x):
            try:
                return not pd.isna(x) and str(x).strip() != "" and int(float(x)) == float(x)
            except Exception:
                return False
            
        # 1. 🔁 Doublons
        df_valid = df[df["Activite"].notna() & (df["Activite"].astype(str).str.strip() != "")]

        # Création d'une colonne temporaire pour la comparaison
        df_valid = df[df["Activite"].notna() & (df["Activite"].astype(str).str.strip() != "")]
        df_valid = df_valid.copy()  # pour éviter SettingWithCopyWarning
        df_valid["_spectacle_clean"] = df_valid["Activite"].astype(str).str.strip().str.lower()
        doublons = df_valid[df_valid.duplicated(subset=["_spectacle_clean"], keep=False)]

        if not doublons.empty:
            bloc = ["🟠 Doublons de spectacle :"]
            for _, row in doublons.iterrows():
                try:
                    date_str = str(int(float(row["Date"]))) if pd.notna(row["Date"]) else "Vide"
                except (ValueError, TypeError):
                    date_str = "Vide"
                heure_str = str(row["Debut"]).strip() if pd.notna(row["Debut"]) else "Vide"
                duree_str = str(row["Duree"]).strip() if pd.notna(row["Duree"]) else "Vide"
                bloc.append(f"{date_str} - {heure_str} - {row['Activite']} ({duree_str})")
            erreurs.append("\n".join(bloc))
            
        # 2. ⛔ Chevauchements
        chevauchements = []
        df_sorted = df.sort_values(by=["Date", "Debut_dt"])
        for i in range(1, len(df_sorted)):
            r1 = df_sorted.iloc[i - 1]
            r2 = df_sorted.iloc[i]
            if r1.isna().all() or r2.isna().all():
                continue
            if pd.notna(r1["Date"]) and pd.notna(r2["Date"]) and r1["Date"] == r2["Date"]:
                fin1 = r1["Debut_dt"] + r1["Duree_dt"]
                debut2 = r2["Debut_dt"]
                if debut2 < fin1:
                    chevauchements.append((r1, r2))
        if chevauchements:
            bloc = ["🔴 Chevauchements:"]
            for r1, r2 in chevauchements:
                bloc.append(
                    f"{r1['Activite']} ({r1['Debut']} / {r1['Duree']}) chevauche {r2['Activite']} ({r2['Debut']} / {r2['Duree']}) le {r1['Date']}"
                )
            erreurs.append("\n".join(bloc))

        # 3. 🕒 Erreurs de format
        bloc_format = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            # Date : uniquement si non NaN
            if pd.notna(row["Date"]) and not est_entier(row["Date"]):
                bloc_format.append(f"Date invalide à la ligne {idx + 2} : {row['Date']}")

            # Ne tester Heure/Duree que si Spectacle ou Autres est renseigné
            if str(row["Activite"]).strip() != "":
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Debut"]).strip()):
                    bloc_format.append(f"Heure invalide à la ligne {idx + 2} : {row['Debut']}")
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Duree"]).strip()):
                    bloc_format.append(f"Durée invalide à la ligne {idx + 2} : {row['Duree']}")
            
            # Test de la colonne Relache
            if not est_relache_valide(row["Relache"]):
                bloc_format.append(f"Relache invalide à la ligne {idx + 2} : {row['Relache']}")

        # 4. 📆 Spectacles un jour de relâche (Date == Relache)
        bloc_relache = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if (
                est_entier(row["Date"]) and
                est_entier(row["Relache"]) and
                int(float(row["Date"])) == int(float(row["Relache"])) and
                str(row["Activite"]).strip() != ""
            ):
                bloc_relache.append(
                    f"{row['Activite']} prévu le jour de relâche ({int(row['Date'])}) à la ligne {idx + 2}"
                )
        if bloc_relache:
            erreurs.append("🛑 Spectacles programmés un jour de relâche:\n" + "\n".join(bloc_relache))

        # 5. 🕳️ Heures non renseignées
        bloc_heure_vide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                if pd.isna(row["Debut"]) or str(row["Debut"]).strip() == "":
                    bloc_heure_vide.append(f"Heure vide à la ligne {idx + 2}")
        if bloc_heure_vide:
            erreurs.append("⚠️ Heures non renseignées:\n" + "\n".join(bloc_heure_vide))

        # 6. 🕓 Heures au format invalide
        bloc_heure_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                h = row["Debut"]
                if pd.notna(h) and str(h).strip() != "":
                    h_str = str(h).strip().lower()
                    is_time_like = isinstance(h, (datetime.datetime, datetime.time))
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", h_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", h_str))
                    if not is_time_like and not valid_format:
                        bloc_heure_invalide.append(f"Heure invalide à la ligne {idx + 2} : {h}")
        if bloc_heure_invalide:
            erreurs.append("⛔ Heures mal formatées:\n" + "\n".join(bloc_heure_invalide))

        # 7. 🕳️ Durées non renseignées ou nulles
        bloc_duree_nulle = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if isinstance(row["Duree_dt"], pd.Timedelta) and row["Duree_dt"] == pd.Timedelta(0):
                if pd.isna(row["Duree"]) or str(row["Duree"]).strip() == "":
                    msg = f"Durée vide à la ligne {idx + 2}"
                else:
                    msg = f"Durée égale à 0 à la ligne {idx + 2} : {row['Duree']}"
                bloc_duree_nulle.append(msg)
        if bloc_duree_nulle:
            erreurs.append("⚠️ Durées nulles ou vides:\n" + "\n".join(bloc_duree_nulle))

        # 8. ⏱️ Durées au format invalide
        bloc_duree_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                d = row["Duree"]
                if pd.notna(d) and str(d).strip() != "":
                    d_str = str(d).strip().lower()
                    is_timedelta = isinstance(d, pd.Timedelta)
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", d_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", d_str))
                    if not is_timedelta and not valid_format:
                        bloc_duree_invalide.append(f"Durée invalide à la ligne {idx + 2} : {d}")
        if bloc_duree_invalide:
            erreurs.append("⛔ Durées mal formatées:\n" + "\n".join(bloc_duree_invalide))

        contenu = "<div style='font-size: 14px;'>"
        for bloc in erreurs:
            lignes = bloc.split("\n")
            if lignes[0].startswith(("🟠", "🔴", "⚠️", "🛑", "⛔")):
                contenu += f"<p><strong>{lignes[0]}</strong></p><ul>"
                for ligne in lignes[1:]:
                    contenu += f"<li>{ligne}</li>"
                contenu += "</ul>"
            else:
                contenu += f"<p>{bloc}</p>"
        contenu += "</div>"
        return contenu

    with st.expander("Cohérence des données"):
        st.markdown(get_log_verifier_coherence(df), unsafe_allow_html=True)

# Indique si une row est une activité programmée
def est_activite_programmee(row):
    return (est_float_valide(row["Date"]) and 
             pd.notna(row["Debut"]) and 
             pd.notna(row["Duree"]) and 
             pd.notna(row["Activite"]))


# Renvoie le dataframe des activités programmées
def get_activites_programmees(df):
    return df[
        df["Date"].apply(est_float_valide) & 
        df["Debut"].notna() & 
        df["Duree"].notna() &
        df["Activite"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Renvoie le dataframe des activités non programmées
def get_activites_non_programmees(df):
    return df[df["Date"].isna() & 
              df["Activite"].notna() & 
              df["Debut"].notna() & 
              df["Fin"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])
    # return df[~(
    #     df["Date"].apply(est_float_valide) & 
    #     df["Debut"].notna() & 
    #     df["Duree"].notna() &
    #     df["Activite"].notna()
    # )].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Affiche le bouton de recharche sur le web
def afficher_bouton_web(nom_activite, disabled=False):    

    #Retour si nom activité vide
    if pd.isna(nom_activite):
        return
                
    # Initialiser le dictionnaire si nécessaire
    if "liens_activites" not in st.session_state:
        st.session_state["liens_activites"] = {}

    liens = st.session_state["liens_activites"]

    # Vérifier si un lien existe déjà
    if nom_activite in liens:
        url = liens[nom_activite]
    else:
        # Construire l'URL de recherche
        url = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={nom_activite.replace(' ', '+')}"
        if nom_activite in liens:
            liens[nom_activite] = url  # L'enregistrer dans la session

    st.link_button(LABEL_BOUTON_CHERCHER_WEB, url, use_container_width=CENTRER_BOUTONS, disabled=disabled)

# Détection basique de plateforme
def get_platform():
    if "platform" in st.session_state:
        return st.session_state["platform"]

    user_agent = st_javascript("navigator.userAgent", key="user_agent_detect")
    if user_agent == 0 or user_agent is None:
        debug_trace("Détection plateforme")
        st.stop()

    # Traitement une fois la valeur reçue
    ua = user_agent.lower()
    if "iphone" in ua or "ipad" in ua or "ipod" in ua:
        platform = "iOS"
    elif "android" in ua:
        platform = "Android"
    elif "windows" in ua:
        platform = "Windows"
    elif "macintosh" in ua:
        platform = "macOS"
    elif "linux" in ua:
        platform = "Linux"
    else:
        platform = "Autre"

    debug_trace("Plateforme détectée")

    st.session_state["platform"] = platform
    st.rerun()   

from difflib import SequenceMatcher

def _normalize(txt: str) -> str:
    if not isinstance(txt, str):
        return ""
    # minuscules + sans accents + espaces compactés
    t = unicodedata.normalize("NFD", txt).encode("ascii", "ignore").decode("ascii")
    t = re.sub(r"\s+", " ", t.strip().lower())
    return t

def _best_match_row(carnet_df: pd.DataFrame, key_norm: str):
    """
    Retourne l'index de la meilleure ligne matchée dans carnet_df
    selon l'ordre: égalité stricte > contains > fuzzy.
    Renvoie None si aucun candidat crédible.
    """
    if carnet_df.empty:
        return None

    # Prépare colonne normalisée
    if "_Nom_norm" not in carnet_df.columns:
        carnet_df["_Nom_norm"] = carnet_df["Nom"].astype(str).apply(_normalize)

    noms = carnet_df["_Nom_norm"]

    # 1) égalité stricte
    exact = carnet_df.index[noms == key_norm]
    if len(exact):
        return exact[0]

    # 2) contains (key dans nom)
    contains_idx = [i for i, n in noms.items() if key_norm in n]
    if contains_idx:
        # si plusieurs, prend le plus proche via ratio fuzzy
        if len(contains_idx) == 1:
            return contains_idx[0]
        best = max(contains_idx, key=lambda i: SequenceMatcher(None, key_norm, noms[i]).ratio())
        return best

    # 3) fuzzy global (utile si fautes de frappe)
    # on prend les candidats avec ratio >= 0.75 et choisit le meilleur
    scored = [(i, SequenceMatcher(None, key_norm, n).ratio()) for i, n in noms.items()]
    scored = [x for x in scored if x[1] >= 0.75]
    if scored:
        scored.sort(key=lambda x: x[1], reverse=True)
        return scored[0][0]

    return None

@st.cache_data(show_spinner=False)
def prepare_carnet(carnet_df: pd.DataFrame) -> pd.DataFrame:
    """Ajoute une colonne normalisée (une seule fois, puis cache)."""
    df = carnet_df.copy()
    if "Nom" in df.columns:
        df["_Nom_norm"] = df["Nom"].astype(str).map(_normalize)
    else:
        df["_Nom_norm"] = ""
    return df

def resolve_address_fast(lieu: str, carnet_df: pd.DataFrame | None, city_default="Avignon"):
    """
    1) Cherche dans le carnet par égalité puis 'contains' (normalisé, sans accents).
    2) Si rien -> renvoie 'lieu, <city>'.
    Retourne (addr_humaine, addr_enc).
    """
    lieu = lieu if isinstance(lieu, str) else ""
    lieu = lieu.strip()
    key = _normalize(lieu)

    addr = ""
    if carnet_df is not None and {"Nom","Adresse"}.issubset(carnet_df.columns):
        df = prepare_carnet(carnet_df)

        # match exact (rapide)
        hit = df.loc[df["_Nom_norm"].eq(key)]
        if hit.empty and key:
            # contains (vectorisé)
            hit = df.loc[df["_Nom_norm"].str.contains(re.escape(key), na=False)]

        if not hit.empty:
            val = hit.iloc[0]["Adresse"]
            if pd.notna(val):
                addr = str(val).strip()

    if not addr:
        # fallback toujours: lieu + ville
        addr = f"{lieu}, {city_default}" if lieu else city_default

    return addr, quote_plus(addr)

def resolve_address(lieu: str, carnet_df: pd.DataFrame | None = None, default_city="Avignon"):
    """
    Retourne (addr_humaine, addr_enc) en essayant d'abord le carnet (Nom -> Adresse)
    avec recherche accent-insensible, partielle, et fuzzy.
    Si pas trouvé, ajoute toujours ", <city>" au lieu.
    """
    lieu = lieu if isinstance(lieu, str) else ""
    saisie = lieu.strip()
    key = _normalize(saisie)

    addr = ""

    if carnet_df is not None and {"Nom", "Adresse"}.issubset(carnet_df.columns):
        try:
            row_idx = _best_match_row(carnet_df, key)
            if row_idx is not None:
                val = carnet_df.loc[row_idx, "Adresse"]
                if pd.notna(val):
                    addr = str(val).strip()
        except Exception:
            pass  # pas de blocage si carnet mal formé

    # Fallback : toujours ajouter la ville si rien trouvé
    if not addr:
        if saisie:
            addr = f"{saisie}, {default_city}"
        else:
            addr = default_city

    addr_enc = quote_plus(addr) if addr else ""
    return addr, addr_enc

# Affiche le bouton de recherche d'itinéraire
def afficher_bouton_itineraire(lieu, disabled=False):  

    # Bouton désactivé si lieu vide ou None
    if pd.isna(lieu) or not str(lieu).strip():
        st.link_button(
            LABEL_BOUTON_CHERCHER_ITINERAIRE,
            "#",  # pas de lien cliquable
            use_container_width=CENTRER_BOUTONS,
            disabled=True
        )
        return
    
     # Résolution depuis carnet + fallback
    addr_human, addr_enc = resolve_address_fast(lieu, st.session_state.carnet_adresses, city_default=st.session_state.city_default)
    itineraire_app = st.session_state.get("itineraire_app", "Google Maps Web")
    platform = get_platform()  

    if itineraire_app == "Apple Maps" and platform == "iOS":
        url = f"http://maps.apple.com/?daddr={addr_enc}"

    elif itineraire_app == "Google Maps App":
        if platform == "iOS":
            url = f"comgooglemaps://?daddr={addr_enc}"
        elif platform == "Android":
            url = f"geo:0,0?q={addr_enc}"
        else:
            # Sur desktop, on retombe sur la version web
            url = f"https://www.google.com/maps/dir/?api=1&destination={addr_enc}"

    else:  # Google Maps Web
        url = f"https://www.google.com/maps/dir/?api=1&destination={addr_enc}"

    st.link_button(
        LABEL_BOUTON_CHERCHER_ITINERAIRE,
        url,
        use_container_width=CENTRER_BOUTONS,
        disabled=disabled or not addr_enc
    )

# Indique si une activité donnée par son descripteur dans le df est réservée
def est_reserve(ligne_df):
    return str(ligne_df["Reserve"]).strip().lower() == "oui"

# Renvoie les lignes modifées entre df1 et df2, l'index de df2 est supposé se trouver dans la colonne __index de df1
def get_lignes_modifiees(df1, df2, columns_to_drop=[]):
    lignes_modifiees = set()
    for i, row in df1.iterrows():
        idx = row["__index"]
        for col in df1.drop(columns=columns_to_drop).columns:
            if idx in df2.index:
                val_avant = df2.at[idx, col]
                val_apres = row[col]
                if pd.isna(val_avant) and pd.isna(val_apres):
                    continue
                if (pd.isna(val_avant) and pd.notna(val_apres)) or val_avant != val_apres:
                    lignes_modifiees.add((i, idx))
    return lignes_modifiees

# DialogBox de suppression d'activité
@st.dialog("Suppression activité")
def show_dialog_supprimer_activite(df, index_df, df_display):
    st.markdown("Voulez-vous supprimer cette activité ?")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            undo_redo_save()
            if est_activite_programmee(df.loc[index_df]):
                st.session_state.activites_programmees_selected_row = ligne_voisine_index(df_display, index_df)
                forcer_reaffichage_activites_programmees()
            else:
                st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(df_display, index_df)
                forcer_reaffichage_activites_non_programmees()
            forcer_reaffichage_df("creneaux_disponibles")
            supprimer_activite(df, index_df)
            sauvegarder_row_ds_gsheet(df, index_df)
            st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# DialogBox de reprogrammation d'activité programmée
@st.dialog("Reprogrammation activité")
def show_dialog_reprogrammer_activite_programmee(df, index_df, df_display, jours_possibles):
    jour_escape = "Aucune" # escape pour déprogrammer l'activité
    jours_possibles = get_jours_possibles(df, get_activites_programmees(df), index_df) + [jour_escape]
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles[:-1]] + [jours_possibles[-1]]
    jour_selection = st.selectbox("Choisissez une nouvelle date pour cette activité :", jours_label, key = "ChoixJourReprogrammationActiviteProgrammee")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            if jour_selection == jour_escape:
                # Suppresion de la liste des activités programmées
                undo_redo_save()
                st.session_state.activites_programmees_selected_row = ligne_voisine_index(df_display, index_df)
                st.session_state.activites_non_programmees_selected_row = index_df
                deprogrammer_activite_programmee(df, index_df)
                forcer_reaffichage_activites_programmees()
                forcer_reaffichage_activites_non_programmees()
                forcer_reaffichage_df("creneaux_disponibles")
                sauvegarder_row_ds_gsheet(df, index_df)
                st.rerun()
            else:
                # Reprogrammation 
                jour_choisi = int(jour_selection) 
                undo_redo_save()
                st.session_state.activites_programmees_selected_row = index_df
                df.at[index_df, "Date"] = jour_choisi
                forcer_reaffichage_activites_programmees()
                sauvegarder_row_ds_gsheet(df, index_df)
                st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# DialogBox de programmation d'activité non programmée
@st.dialog("Programmation activité")
def show_dialog_reprogrammer_activite_non_programmee(df, index_df, df_display, jours_possibles):
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    jour_selection = st.selectbox("Choisissez une date pour cette activité :", jours_label, key = "ChoixJourProgrammationActiviteNonProgrammee")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            # Programmation à la date choisie
            jour_choisi = int(jour_selection.split()[-1])
            undo_redo_save()
            st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(df_display, index_df)
            st.session_state.activites_programmees_selected_row = index_df
            df.at[index_df, "Date"] = jour_choisi
            forcer_reaffichage_activites_programmees()
            forcer_reaffichage_activites_non_programmees()
            forcer_reaffichage_df("creneaux_disponibles")
            sauvegarder_row_ds_gsheet(df, index_df)
            st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# Affiche les activités programmées dans un tableau
def afficher_activites_programmees(df):

    st.markdown("##### Activités programmées")

    # Constitution du df à afficher
    activites_programmees = get_activites_programmees(df)
    st.session_state.activites_programmees = activites_programmees
    df_display = activites_programmees.copy()
    df_display["__jour"] = df_display["Date"].apply(lambda x: int(str(int(float(x)))[-2:]) if pd.notna(x) else None)
    df_display["__index"] = df_display.index
    df_display["__options_date"] = calculer_options_date_activites_programmees(df_display) 
    df_display["__non_reserve"] = df_display["Reserve"].astype(str).str.strip().str.lower() != "oui"
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df_display)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 250
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de savoir si l'on doit forcer le réaffichage de l'aggrid après une suppression de ligne 
    if "aggrid_activites_programmees_reset_counter" not in st.session_state:
        st.session_state.aggrid_activites_programmees_reset_counter = 0

    # Initialisation du flag permettant de savoir si l'on est en mode réaffichage complet de l'aggrid
    if "aggrid_activites_programmees_forcer_reaffichage" not in st.session_state:
        st.session_state.aggrid_activites_programmees_forcer_reaffichage = False
   
    # Initialisation du flag permettant de savoir si l'on doit gérer les modifications de cellules
    if "aggrid_activites_programmees_gerer_modification_cellule" not in st.session_state:
        st.session_state.aggrid_activites_programmees_gerer_modification_cellule = True
   
    # Initialisation de la variable d'état contenant l'index de ligne sélectionnée courant
    if "activites_programmees_selected_row" not in st.session_state:
        st.session_state.activites_programmees_selected_row = None
   
    # Enregistrement dans st.session_state d'une copy du df à afficher
    st.session_state.activites_programmees_df_display = df_display.copy()

    # Configuration
    gb = GridOptionsBuilder.from_dataframe(df_display)

    # Masquage des colonnes de travail
    work_cols = ["__index", "__jour", "__options_date", "__non_reserve"]
    for col in work_cols:
        gb.configure_column(col, hide=True)

    # Colonnes editables
    non_editable_cols = ["Fin"] + work_cols
    for col in df_display.columns:
        gb.configure_column(col, editable=(col not in non_editable_cols))

    gb.configure_column(
        "Début",
        editable=JsCode("function(params) { return params.data.__non_reserve; }")
    )

    gb.configure_column(
        "Durée" \
        "",
        editable=JsCode("function(params) { return params.data.__non_reserve; }")
    )

    # Configuration des menus de la colonne Date
    gb.configure_column(
        "Date",
        editable=True,
        cellEditor="agSelectCellEditor",
        cellEditorParams=JsCode("""
            function(params) {
                return {
                    values: params.data.__options_date || []
                }
            }
        """),
    )

    # Colorisation
    gb.configure_grid_options(getRowStyle=JsCode(f"""
    function(params) {{
        const jour = params.data.__jour;
        const couleurs = {PALETTE_COULEURS_JOURS};
        if (jour && couleurs[jour]) {{
            return {{ 'backgroundColor': couleurs[jour] }};
        }}
        return null;
    }}
    """))

    # Retaillage largeur colonnes
    gb.configure_default_column(resizable=True)

    # Configuration de la sélection
    pre_selected_row = 0  # par défaut
    if "activites_programmees_selected_row" in st.session_state:
        valeur_index = st.session_state["activites_programmees_selected_row"]
        matches = df_display[df_display["__index"].astype(str) == str(valeur_index)]
        if not matches.empty:
            pre_selected_row = df_display.index.get_loc(matches.index[0])
    gb.configure_selection(selection_mode="single", use_checkbox=False, pre_selected_rows=[pre_selected_row])
    
    js_code = JsCode(f"""
            function(params) {{
                params.api.sizeColumnsToFit();
                params.api.ensureIndexVisible({pre_selected_row}, 'middle');
                params.api.getDisplayedRowAtIndex({pre_selected_row}).setSelected(true);
            }}
        """)
    gb.configure_grid_options(onGridReady=js_code)

    grid_options = gb.build()
    grid_options["suppressMovableColumns"] = True

    # Affectation du masque d'évènements UI à prendre en compte
    if "activites_programmees_update_on" not in st.session_state:
        if MENU_ACTIVITE_UNIQUE:
            st.session_state.activites_programmees_update_on = AGGRID_UPDATE_ON_MENU_ACTIVITE_UNIQUE  
        else:
            st.session_state.activites_programmees_update_on = AGGRID_UPDATE_ON_MENU_ACTIVITE_DOUBLE

    # Affichage
    response = AgGrid(
        df_display,
        gridOptions=grid_options,
        allow_unsafe_jscode=True,
        height=height,
        update_on=st.session_state.activites_programmees_update_on,
        key=f"Activités programmées {st.session_state.aggrid_activites_programmees_reset_counter}",  # clé stable mais changeante après suppression de ligne pour forcer le reaffichage
    )

    event_data = response.get("event_data")
    event_type = event_data["type"] if isinstance(event_data, dict) else None

    # Pas d'event aggrid à traiter si event_type is None (i.e. le script python est appelé pour autre chose qu'un event aggrid)
    if event_type is None:
        if not st.session_state.get("sidebar_menus"):
            if len(df_display) == 0:
                with st.expander("Contrôles"):
                    menu_activites_programmees(df, None, df_display, "")
        return

    debug_trace(f"PROG {event_type}", trace_type=["gen", "event"])

    # Récupération de la ligne sélectionnée courante
    selected_rows = response["selected_rows"]
    if st.session_state.aggrid_activites_programmees_forcer_reaffichage == True:
        row = df_display.iloc[pre_selected_row] if pre_selected_row < len(df_display) else pd.Series({col: np.nan for col in df_display.columns})
    else:
        if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
            row = selected_rows.iloc[0] 
        elif isinstance(selected_rows, list) and len(selected_rows) > 0:
            row = selected_rows[0]
        else: 
            row = df_display.iloc[pre_selected_row] if pre_selected_row < len(df_display) else pd.Series({col: np.nan for col in df_display.columns})
    st.session_state.aggrid_activites_programmees_forcer_reaffichage = False

    # 🟡 Traitement si ligne sélectionnée et index correspondant non vide
    if row is not None and pd.notna(row["__index"]):

        # Récupération de l'index de ligne sélectionnée
        index_df = row["__index"]

        # Recupération du nom d'activité correspondant
        nom_activite = str(row["Activité"]).strip() if pd.notna(row["Activité"]) else ""

        # Gestion du menu activité en fonction du type d'évènement renvoyé
        event_data = response.get("event_data")
        event_type = event_data["type"] if isinstance(event_data, dict) else None

        if MENU_ACTIVITE_UNIQUE:
            # La détection des "headerFocused" est spécifique à chaque grille à la différence des rowClicked et cellClicked  
            # et permet de mettre à jour le menu activité de maniere alternative au selectionChanged qui ne se déclenche que sur 
            # changement de sélection donc pas si l'on clique la row déjà sélectionnée
            if event_type == "headerFocused": 

                # Mise à jour du menu activité géré par afficher_sidebar()
                if st.session_state.forcage_menu_activites_enabled and st.session_state.forcer_menu_activites_non_programmees:
                    st.session_state.forcer_menu_activites_non_programmees = False
                else:
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_programmees",
                        "df": df,
                        "index_df": index_df,
                        "df_display": df_display,
                        "nom_activite": nom_activite
                    }

            # Evènement de type "selectionChanged" 
            elif event_type == "selectionChanged":
                if index_df != st.session_state.activites_programmees_selected_row:
                    st.session_state.activites_programmees_selected_row = index_df
                    st.session_state.editeur_activite_courante_idx = index_df
                    
                    # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

                    # Mise à jour du menu activité géré par afficher_sidebar()
                    if st.session_state.forcage_menu_activites_enabled and st.session_state.forcer_menu_activites_non_programmees:
                        st.session_state.forcer_menu_activites_non_programmees = False
                    else:
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_programmees",
                            "df": df,
                            "index_df": index_df,
                            "df_display": df_display,
                            "nom_activite": nom_activite
                        }

        else:
            if index_df != st.session_state.activites_programmees_selected_row:
                st.session_state.activites_programmees_selected_row = index_df
                st.session_state.editeur_activite_courante_idx = index_df

            # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

            # Mise à jour du menu activité géré par afficher_sidebar()
            st.session_state.menu_activites_programmees = {
                "menu": "menu_activites_programmees",
                "df": df,
                "index_df": index_df,
                "df_display": df_display,
                "nom_activite": nom_activite
            }
                    
        # Affichage de l'erreur renvoyée par le précédent run
        erreur = st.session_state.get("aggrid_activites_programmees_erreur") 
        if erreur is not None:
            st.error(erreur)

        # Gestion des modifications de cellules
        if st.session_state.aggrid_activites_programmees_gerer_modification_cellule == True:
            if isinstance(response["data"], pd.DataFrame):
                df_modifie = pd.DataFrame(response["data"])
                lignes_modifiees = get_lignes_modifiees(df_modifie, st.session_state.activites_programmees_df_display, columns_to_drop=work_cols)
                if lignes_modifiees:
                    st.session_state.aggrid_activites_programmees_erreur = None
                    undo_redo_save()
                    for i, idx in lignes_modifiees:
                        for col in [col for col in df_modifie.columns if col not in non_editable_cols]:
                            col_df = RENOMMAGE_COLONNES_INVERSE[col] if col in RENOMMAGE_COLONNES_INVERSE else col
                            if pd.isna(df.at[idx, col_df]) and pd.isna(df_modifie.at[i, col]):
                                continue
                            if col == "Date":
                                if df_modifie.at[i, col] == "":
                                    # Déprogrammation de l'activité (Suppression de l'activité des activités programmées)
                                    st.info("Déprogrammation de l'activité")
                                    undo_redo_save()
                                    st.session_state.activites_programmees_selected_row = ligne_voisine_index(df_display, idx)
                                    st.session_state.activites_non_programmees_selected_row = idx
                                    deprogrammer_activite_programmee(df, idx)
                                    forcer_reaffichage_activites_programmees()
                                    forcer_reaffichage_activites_non_programmees()
                                    forcer_reaffichage_df("creneaux_disponibles")
                                    sauvegarder_row_ds_gsheet(df, idx)
                                    st.rerun()
                                elif df_modifie.at[i, col] != str(int(df.at[idx, "Date"])):
                                    # Reprogrammation de l'activité à la date choisie
                                    st.info("Reprogrammation de l'activité à la date choisie")
                                    jour_choisi = int(df_modifie.at[i, col])
                                    undo_redo_save()
                                    st.session_state.activites_programmees_selected_row = idx
                                    df.at[idx, "Date"] = jour_choisi
                                    forcer_reaffichage_activites_programmees()
                                    sauvegarder_row_ds_gsheet(df, idx)
                                    st.rerun()
                                else:
                                    st.info("Activité réservée")
                            else:
                                if (pd.isna(df.at[idx, col_df]) and pd.notna(df_modifie.at[i, col])) or df.at[idx, col_df] != df_modifie.at[i, col]:
                                    # st.session_state.editeur_activites_programmees_utiliser_index_colonne_courante = True
                                    erreur = affecter_valeur(df,idx, col_df, df_modifie.at[i, col])
                                    if not erreur:
                                        forcer_reaffichage_activites_programmees()
                                        if col in ["Debut", "Duree", "Activité"]:
                                            forcer_reaffichage_df("creneaux_disponibles")
                                        st.rerun()
                                    else:
                                        st.session_state.aggrid_activites_programmees_erreur = erreur
                                        forcer_reaffichage_activites_programmees()
                                        st.rerun()
        st.session_state.aggrid_activites_programmees_gerer_modification_cellule = True

        if not st.session_state.get("sidebar_menus"):
            with st.expander("Contrôles"):
                menu_activites_programmees(df, index_df, df_display, nom_activite)

    elif not MENU_ACTIVITE_UNIQUE:
        if len(df_display) == 0:
            st.session_state.menu_activites_programmees = {
                "menu": "menu_activites_programmees",
                "df": df,
                "index_df": None,
                "df_display": df_display,
                "nom_activite": ""
            }
    
# Menu activité à afficher dans la sidebar si click dans aggrid d'activités programmées         }
def menu_activites_programmees(df, index_df, df_display, nom_activite):

    boutons_disabled = nom_activite == "" or pd.isna(index_df) or not isinstance(df, pd.DataFrame) or (isinstance(df, pd.DataFrame) and len(df) == 0)
    activite_reservee = est_reserve(df.loc[index_df]) if pd.notna(index_df) else True 
    jours_possibles = get_jours_possibles(df, get_activites_programmees(df), index_df)

    # Affichage du label d'activité
    if MENU_ACTIVITE_UNIQUE:
        afficher_nom_activite_clickable(df, index_df, nom_activite)
    else:
        afficher_nom_activite(df, index_df, nom_activite)

    # Affichage du contrôle recherche sur le Web
    afficher_bouton_web(nom_activite, disabled=boutons_disabled or est_pause_str(nom_activite))

    # Affichage du contrôle recherche itinéraire
    afficher_bouton_itineraire(df.loc[index_df, "Lieu"] if pd.notna(index_df) and len(df) > 0 else "")

    # Affichage contrôle Supprimer
    if st.button(LABEL_BOUTON_SUPPRIMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or activite_reservee, key="SupprimerActiviteProgrammee"):
        undo_redo_save()
        st.session_state.activites_programmees_selected_row = ligne_voisine_index(df_display, index_df)
        supprimer_activite(df, index_df)
        forcer_reaffichage_activites_programmees()
        sauvegarder_row_ds_gsheet(df, index_df)
        st.rerun()

    # Affichage contrôle Deprogrammer
    if st.button(LABEL_BOUTON_DEPROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or activite_reservee, key="DéprogrammerActivitéProgrammee"):
        undo_redo_save()
        st.session_state.activites_programmees_selected_row = ligne_voisine_index(df_display, index_df)
        st.session_state.activites_non_programmees_selected_row = index_df

        if MENU_ACTIVITE_UNIQUE:
            # Bascule du menu activité sur le menu_activites_non_programmees
            st.session_state.forcage_menu_activites_enabled = True
            st.session_state.forcer_menu_activites_non_programmees = True
            st.session_state.menu_activites = {
                "menu": "menu_activites_non_programmees",
                "df": df,
                "index_df": index_df,
                "df_display": st.session_state.activites_non_programmees_df_display,
                "nom_activite": nom_activite
            }

        deprogrammer_activite_programmee(df, index_df)
        forcer_reaffichage_activites_programmees()
        forcer_reaffichage_activites_non_programmees()
        forcer_reaffichage_df("creneaux_disponibles")
        sauvegarder_row_ds_gsheet(df, index_df)
        st.rerun()
    else:
        if MENU_ACTIVITE_UNIQUE:
            if st.session_state.forcage_menu_activites_enabled:
                if st.session_state.activites_programmees_selected_row == st.session_state.menu_activites["index_df"]:
                    st.session_state.forcage_menu_activites_enabled = False

    # Affichage contrôle Reprogrammer
    if st.button(LABEL_BOUTON_REPROGRAMMER, use_container_width=True, disabled=boutons_disabled or activite_reservee or not jours_possibles, key="ReprogrammerActivitéProgrammee"):
        if "activites_programmees_jour_choisi" in st.session_state:
            jour_choisi = st.session_state.activites_programmees_jour_choisi
            undo_redo_save()
            st.session_state.activites_programmees_selected_row = index_df
            df.at[index_df, "Date"] = int(jour_choisi)
            forcer_reaffichage_activites_programmees()
            sauvegarder_row_ds_gsheet(df, index_df)
            st.rerun()
    
    # Affichage Liste des jours possibles
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    if jours_label and (not st.session_state.get("ChoixJourReprogrammationActiviteProgrammee") or st.session_state.ChoixJourReprogrammationActiviteProgrammee not in jours_label):
            st.session_state.ChoixJourReprogrammationActiviteProgrammee = jours_label[0]
    st.session_state.activites_programmees_jour_choisi = st.selectbox("Jours possibles", jours_label, label_visibility="visible", disabled=boutons_disabled or activite_reservee or not jours_possibles, key = "ChoixJourReprogrammationActiviteProgrammee") 
        
    # Affichage de l'éditeur d'activité
    if st.button(LABEL_BOUTON_EDITER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled, key="menu_activites_programmees_bouton_editeur"):
        show_dialog_editeur_activite(df, index_df)
                               
    if  MENU_ACTIVITE_UNIQUE and st.session_state.sidebar_menus:
        # Affichage contrôle Ajouter
        afficher_bouton_nouvelle_activite(df, key="ajouter_activites_programmees")
    
# Affiche les activités non programmées dans un tableau
def afficher_activites_non_programmees(df):

    st.markdown("##### Activités non programmées")

    # Constitution du df à afficher
    df_display = get_activites_non_programmees(df)
    df_display["__index"] = df_display.index
    df_display["__options_date"] = calculer_options_date_activites_non_programmees(df_display) 
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df_display)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 250
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de savoir si l'on doit forcer le réaffichage de l'aggrid après une suppression de ligne 
    if "aggrid_activites_non_programmees_reset_counter" not in st.session_state:
        st.session_state.aggrid_activites_non_programmees_reset_counter = 0
    
    # Initialisation du flag permettant de savoir si l'on est en mode réaffichage complet de l'aggrid
    if "aggrid_activites_non_programmees_forcer_reaffichage" not in st.session_state:
        st.session_state.aggrid_activites_non_programmees_forcer_reaffichage = False
   
    # Initialisation du flag permettant de savoir si l'on doit gérer les modifications de cellules
    if "aggrid_activites_non_programmees_gerer_modification_cellule" not in st.session_state:
        st.session_state.aggrid_activites_non_programmees_gerer_modification_cellule = True
   
    # Initialisation de la variable d'état contenant l'index de ligne sélectionnée courant
    if "activites_non_programmees_selected_row" not in st.session_state:
        st.session_state.activites_non_programmees_selected_row = None
        
    # Enregistrement dans st.session_state d'une copy du df à afficher
    st.session_state.activites_non_programmees_df_display = df_display.copy()

    # Configuration
    gb = GridOptionsBuilder.from_dataframe(df_display)

    # Masquage des colonnes de travail
    work_cols = ["__index", "__options_date"]
    for col in work_cols:
        gb.configure_column(col, hide=True)

    # Colonnes editables
    non_editable_cols = ["Fin"] + work_cols
    for col in df_display.columns:
        gb.configure_column(col, editable=(col not in non_editable_cols))

    # Configuration des menus de la colonne Date
    gb.configure_column(
        "Date",
        editable=True,
        cellEditor="agSelectCellEditor",
        cellEditorParams=JsCode("""
            function(params) {
                return {
                    values: params.data.__options_date || []
                }
            }
        """),
    )

    # Retaillage largeur colonnes
    gb.configure_default_column(resizable=True)

    # Configuration de la sélection
    pre_selected_row = 0  # par défaut
    if "activites_non_programmees_selected_row" in st.session_state:
        valeur_index = st.session_state["activites_non_programmees_selected_row"]
        matches = df_display[df_display["__index"].astype(str) == str(valeur_index)]
        if not matches.empty:
            pre_selected_row = df_display.index.get_loc(matches.index[0])
    gb.configure_selection(selection_mode="single", use_checkbox=False, pre_selected_rows=[pre_selected_row])
    
    js_code = JsCode(f"""
            function(params) {{
                params.api.sizeColumnsToFit();
                params.api.ensureIndexVisible({pre_selected_row}, 'middle');
                params.api.getDisplayedRowAtIndex({pre_selected_row}).setSelected(true);
            }}
        """)
    gb.configure_grid_options(onGridReady=js_code)

    grid_options = gb.build()
    grid_options["suppressMovableColumns"] = True

    # Affectation du masque d'évènements UI à prendre en compte
    if "activites_non_programmees_update_on" not in st.session_state:
        if MENU_ACTIVITE_UNIQUE:
            st.session_state.activites_non_programmees_update_on = AGGRID_UPDATE_ON_MENU_ACTIVITE_UNIQUE  
        else:
            st.session_state.activites_non_programmees_update_on = AGGRID_UPDATE_ON_MENU_ACTIVITE_DOUBLE

    # Affichage
    response = AgGrid(
        df_display,
        gridOptions=grid_options,
        allow_unsafe_jscode=True,
        height=height,
        update_on=st.session_state.activites_non_programmees_update_on,
        key=f"Activités non programmées {st.session_state.aggrid_activites_non_programmees_reset_counter}",  # clé stable mais changeante après suppression de ligne ou modification de cellule pour forcer le reaffichage
    )

    event_data = response.get("event_data")
    event_type = event_data["type"] if isinstance(event_data, dict) else None

    # Pas d'event aggrid à traiter si event_type is None (i.e. le script python est appelé pour autre chose qu'un event aggrid)
    if event_type is None:
        if not st.session_state.sidebar_menus:
            if len(df_display) == 0:
                with st.expander("Contrôles"):
                    menu_activites_non_programmees(df, None, df_display, "")
        return
    
    debug_trace(f"NONPROG {event_type}", trace_type=["gen", "event"])

    # Récupération de la ligne sélectionnée
    selected_rows = response["selected_rows"]
    if st.session_state.aggrid_activites_non_programmees_forcer_reaffichage == True:
        row = df_display.iloc[pre_selected_row] if pre_selected_row < len(df_display) else pd.Series({col: np.nan for col in df_display.columns})
    else:
        if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
            row = selected_rows.iloc[0] 
        elif isinstance(selected_rows, list) and len(selected_rows) > 0:
            row = selected_rows[0]
        else: 
            row = df_display.iloc[pre_selected_row] if pre_selected_row < len(df_display) else pd.Series({col: np.nan for col in df_display.columns})
    st.session_state.aggrid_activites_non_programmees_forcer_reaffichage = False

    # 🟡 Traitement si ligne sélectionnée et index correspondant non vide
    if row is not None and pd.notna(row["__index"]):

        # Récupération de l'index de ligne sélectionnée
        index_df = row["__index"]

        # Recupération du nom d'activité correspondant
        nom_activite = str(row["Activité"]).strip() if pd.notna(row["Activité"]) else ""

        # Gestion du menu activité en fonction du type d'évènement renvoyé
        event_data = response.get("event_data")
        event_type = event_data["type"] if isinstance(event_data, dict) else None

        if MENU_ACTIVITE_UNIQUE:
            # La détection des "headerFocused" est spécifique à chaque grille à la différence des rowClicked et cellClicked  
            # et permet de mettre à jour le menu activité de maniere alternative au selectionChanged qui ne se déclenche que sur 
            # changement de sélection donc pas si l'on clique la row déjà sélectionnée
            if event_type == "headerFocused": 

                # Mise à jour du menu activité géré par afficher_sidebar()
                if st.session_state.forcage_menu_activites_enabled and st.session_state.forcer_menu_activites_programmees:
                    st.session_state.forcer_menu_activites_programmees = False
                else:
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_non_programmees",
                        "df": df,
                        "index_df": index_df,
                        "df_display": df_display,
                        "nom_activite": nom_activite
                    }

            # Evènement de type "selectionChanged"
            elif event_type == "selectionChanged":
                if index_df != st.session_state.activites_non_programmees_selected_row:
                    st.session_state.activites_non_programmees_selected_row = index_df
                    st.session_state.editeur_activite_courante_idx = index_df

                    # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

                    # Mise à jour du menu activité géré par afficher_sidebar()
                    if st.session_state.forcage_menu_activites_enabled and st.session_state.forcer_menu_activites_programmees:
                        st.session_state.forcer_menu_activites_programmees = False
                    else:
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_non_programmees",
                            "df": df,
                            "index_df": index_df,
                            "df_display": df_display,
                            "nom_activite": nom_activite
                        }

        else:
            if index_df != st.session_state.activites_non_programmees_selected_row:
                st.session_state.activites_non_programmees_selected_row = index_df
                st.session_state.editeur_activite_courante_idx = index_df

            # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

            # Mise à jour du menu activité géré par afficher_sidebar()
            st.session_state.menu_activites_non_programmees = {
                "menu": "menu_activites_non_programmees",
                "df": df,
                "index_df": index_df,
                "df_display": df_display,
                "nom_activite": nom_activite
            }

        # Affichage de l'erreur renvoyée par le précédent run
        erreur = st.session_state.get("aggrid_activites_non_programmees_erreur") 
        if erreur is not None:
            st.error(erreur)

        # Gestion des modifications de cellules
        if st.session_state.aggrid_activites_non_programmees_gerer_modification_cellule == True:
            if isinstance(response["data"], pd.DataFrame):
                df_modifie = pd.DataFrame(response["data"])
                lignes_modifiees = get_lignes_modifiees(df_modifie, st.session_state.activites_non_programmees_df_display, columns_to_drop=work_cols)
                if lignes_modifiees:
                    undo_redo_save()
                    st.session_state.aggrid_activites_non_programmees_erreur = None
                    for i, idx in lignes_modifiees:
                        for col in [col for col in df_modifie.columns if col not in non_editable_cols]:
                            col_df = RENOMMAGE_COLONNES_INVERSE[col] if col in RENOMMAGE_COLONNES_INVERSE else col
                            if pd.isna(df.at[idx, col_df]) and pd.isna(df_modifie.at[i, col]):
                                continue
                            if col == "Date":
                                if df_modifie.at[i, col] != "":
                                    # Programmation de l'activité à la date choisie
                                    st.info("Programmation de l'activité à la date choisie")
                                    jour_choisi = int(df_modifie.at[i, col])
                                    undo_redo_save()
                                    st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(df_display, idx)
                                    st.session_state.activites_programmees_selected_row = idx
                                    df.at[idx, "Date"] = jour_choisi
                                    forcer_reaffichage_activites_non_programmees()
                                    forcer_reaffichage_activites_programmees()
                                    sauvegarder_row_ds_gsheet(df, idx)
                                    st.rerun()
                            else:
                                if (pd.isna(df.at[idx, col_df]) and pd.notna(df_modifie.at[i, col])) or df.at[idx, col_df] != df_modifie.at[i, col]:
                                    # st.session_state.editeur_activites_non_programmees_utiliser_index_colonne_courante = True
                                    erreur = affecter_valeur(df,idx, col_df, df_modifie.at[i, col])
                                    if not erreur:
                                        forcer_reaffichage_activites_non_programmees()
                                        forcer_reaffichage_df("activites_programmables_dans_creneau_selectionne")
                                        st.rerun()
                                    else:
                                        st.session_state.aggrid_activites_non_programmees_erreur = erreur
                                        forcer_reaffichage_activites_non_programmees()
                                        st.rerun()
        st.session_state.aggrid_activites_non_programmees_gerer_modification_cellule = True

        if not st.session_state.get("sidebar_menus"):
            with st.expander("Contrôles"):
                menu_activites_non_programmees(df, index_df, df_display, nom_activite)

    elif len(df_display) == 0:
        if MENU_ACTIVITE_UNIQUE:
            if st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
                st.session_state.menu_activites = {
                    "menu": "menu_activites_non_programmees",
                    "df": df,
                    "index_df": None,
                    "df_display": df_display,
                    "nom_activite": ""
                }
        else:
            st.session_state.menu_activites_non_programmees = {
                "menu": "menu_activites_non_programmees",
                "df": df,
                "index_df": None,
                "df_display": df_display,
                "nom_activite": ""
            }

# Menu activité à afficher dans la sidebar si click dans aggrid d'activités non programmées         }
def menu_activites_non_programmees(df, index_df, df_display, nom_activite):
    
    boutons_disabled = nom_activite == "" or pd.isna(index_df) or not isinstance(df, pd.DataFrame) or (isinstance(df, pd.DataFrame) and len(df) == 0)
    jours_possibles = get_jours_possibles(df, get_activites_programmees(df), index_df)

    # Affichage du label d'activité
    if MENU_ACTIVITE_UNIQUE:
        afficher_nom_activite_clickable(df, index_df, nom_activite)
    else:
        afficher_nom_activite(df, index_df, nom_activite)

    # Affichage du contrôle recherche sur le Web
    afficher_bouton_web(nom_activite, disabled=boutons_disabled or est_pause_str(nom_activite))

    # Affichage du contrôle recherche itinéraire
    afficher_bouton_itineraire(df.loc[index_df, "Lieu"] if pd.notna(index_df) and len(df) > 0 else "")

    # Affichage contrôle Supprimer
    if st.button(LABEL_BOUTON_SUPPRIMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled, key="SupprimerActiviteNonProgrammee"):
        undo_redo_save()
        st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(df_display, index_df)
        supprimer_activite(df, index_df)
        forcer_reaffichage_activites_non_programmees()
        forcer_reaffichage_df("activites_programmable_dans_creneau_selectionne")
        sauvegarder_row_ds_gsheet(df, index_df)
        st.rerun()

    # Affichage contrôle Programmer
    if st.button(LABEL_BOUTON_PROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or not jours_possibles, key="AjouterAuxActivitésProgrammees"):
        if "activites_non_programmees_jour_choisi" in st.session_state:
            jour_choisi = st.session_state.activites_non_programmees_jour_choisi
            undo_redo_save()
            st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(df_display, index_df)
            st.session_state.activites_programmees_selected_row = index_df

            if MENU_ACTIVITE_UNIQUE:
                # Bascule du menu activité sur le menu_activites_programmees
                st.session_state.forcage_menu_activites_enabled = True
                st.session_state.forcer_menu_activites_programmees = True
                st.session_state.menu_activites = {
                    "menu": "menu_activites_programmees",
                    "df": df,
                    "index_df": index_df,
                    "df_display": st.session_state.activites_programmees_df_display,
                    "nom_activite": nom_activite
                }

            df.at[index_df, "Date"] = int(jour_choisi)
            forcer_reaffichage_activites_programmees()
            forcer_reaffichage_activites_non_programmees()
            forcer_reaffichage_df("creneaux_disponibles")
            sauvegarder_row_ds_gsheet(df, index_df)
            st.rerun()
    else:
        if MENU_ACTIVITE_UNIQUE:
            if st.session_state.forcage_menu_activites_enabled:
                if st.session_state.activites_programmees_selected_row == st.session_state.menu_activites["index_df"]:
                    st.session_state.forcage_menu_activites_enabled = False

    # Affichage Liste des jours possibles
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    if jours_label and (not st.session_state.get("ChoixJourProgrammationActiviteNonProgrammee") or st.session_state.ChoixJourProgrammationActiviteNonProgrammee not in jours_label):
            st.session_state.ChoixJourProgrammationActiviteNonProgrammee = jours_label[0]
    st.session_state.activites_non_programmees_jour_choisi = st.selectbox("Jours possibles", jours_label, label_visibility="visible", disabled=boutons_disabled or not jours_possibles, key = "ChoixJourProgrammationActiviteNonProgrammee") # , width=90
        
    # Affichage de l'éditeur d'activité
    if st.button(LABEL_BOUTON_EDITER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled,  key="menu_activites_non_programmees_bouton_editeur"):
        show_dialog_editeur_activite(df, index_df)

    # Affichage contrôle Ajouter
    afficher_bouton_nouvelle_activite(df, key="ajouter_activites_non_programmees")

# Affichage de l'éditeur d'activité en mode modal
@st.dialog("Editeur d'activité")
def show_dialog_editeur_activite(df, index_df):
    afficher_nom_activite(df, index_df, afficher_label=False)
    afficher_editeur_activite(df, index_df, modale=True)

# Affichage de l'éditeur d'activité
def afficher_editeur_activite(df, index_df=None, modale=False, key="editeur_activite"):
    # Rien à faire sur df vide
    if len(df) <= 0:
        return
    
    if index_df is None:
        if "editeur_activite_courante_idx" in st.session_state:
            index_df = st.session_state.editeur_activite_courante_idx 
    
    if index_df is not None:

        row = df.loc[index_df]

        if est_reserve(row):
            colonnes_editables = [col for col in df.columns if col not in ["Date", "Fin", "Debut_dt", "Duree_dt", "Debut", "Duree"]]
        else:
            colonnes_editables = [col for col in df.columns if col not in ["Date", "Fin", "Debut_dt", "Duree_dt"]]

        # Ajout de l'hyperlien aux infos éditables
        colonnes_editables.append("lien Web")

        # Traitement de l'accentuation
        colonnes_editables_avec_accents = [RENOMMAGE_COLONNES.get(col, col) for col in colonnes_editables]
        
        if "editeur_activites_index_colonne_courante" not in st.session_state:
            st.session_state.editeur_activites_index_colonne_courante = 0

        colonne = st.selectbox("⚙️ Colonne", colonnes_editables_avec_accents, key=key+"_selectbox_choix_colonne")
        st.session_state.editeur_activites_index_colonne_courante = colonnes_editables_avec_accents.index(colonne)
        colonne_df = RENOMMAGE_COLONNES_INVERSE[colonne] if colonne in RENOMMAGE_COLONNES_INVERSE else colonne

        valeur_courante = None
        if colonne_df != "lien Web":
            valeur_courante = row[colonne_df]
            if colonne_df in ["Date", "Priorite"]:
                if est_float_valide(valeur_courante):
                    valeur_courante = int(valeur_courante)
        else:
            liens_activites = st.session_state.get("liens_activites")
            if isinstance(liens_activites, dict):
                if row["Activite"] in liens_activites:
                    lien = liens_activites.get(row["Activite"])
                    valeur_courante = lien

        nouvelle_valeur = st.text_input(f"✏️ Valeur", "" if pd.isna(valeur_courante) else str(valeur_courante), key=key+"_valeur") 

        erreur = None
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS, key=key+"_validation"):
                st.session_state.editeur_activite_erreur = None
                if colonne_df == "lien Web":
                    undo_redo_save()
                    if "liens_activites" not in st.session_state:
                        st.session_state.liens_activites = {}
                    liens_activites = st.session_state.liens_activites
                    liens_activites[row["Activite"]] = nouvelle_valeur
                    sauvegarder_lnk_ds_gsheet(liens_activites)
                    st.rerun()
                else:
                    if est_activite_programmee(row):
                        erreur = affecter_valeur(df, index_df, colonne_df, nouvelle_valeur)
                        if not erreur:
                            forcer_reaffichage_activites_programmees()
                            if colonne_df in ["Debut", "Duree", "Activité"]:
                                forcer_reaffichage_df("creneaux_disponibles")
                            st.rerun()
                        else:
                            if not modale:
                                st.session_state.editeur_activite_erreur = erreur
                                forcer_reaffichage_activites_programmees()
                                st.rerun()
                    else:
                        erreur = affecter_valeur(df, index_df, colonne_df, nouvelle_valeur)
                        if not erreur:
                            forcer_reaffichage_activites_non_programmees()
                            forcer_reaffichage_df("activites_programmables_dans_creneau_selectionne")
                            st.rerun()
                        else:
                            if not modale:
                                st.session_state.editeur_activite_erreur = erreur
                                forcer_reaffichage_activites_non_programmees()
                                st.rerun()
        with col2:
            if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
                st.rerun()
        
        if not modale:
            # Affichage de l'erreur renvoyée par le précédent run
            erreur = st.session_state.get("editeur_activite_erreur") 
        if erreur is not None:
            st.error(erreur)
            

# Affecte une nouvelle valeur à une cellule du df donnée par son index et sa colonne
def affecter_valeur(df, index, colonne, nouvelle_valeur, inhiber_gestion_modification_cellule=True):
    valeur_courante = df.at[index, colonne]
    erreur = None
    if colonne == "Debut" and not est_heure_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HHhMM (ex : 10h00)"
    elif colonne == "Duree" and not est_duree_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HhMM (ex : 1h00 ou 0h30)"
    elif colonne == "Relache" and not est_relache_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : 1, 10, pair, impair"
    elif colonne == "Reserve" and not est_reserve_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : Oui, Non"
    elif ptypes.is_numeric_dtype(df[colonne]) and not ptypes.is_numeric_dtype(nouvelle_valeur):
        try:
            if "." not in nouvelle_valeur and "," not in nouvelle_valeur and "e" not in nouvelle_valeur.lower():
                nouvelle_valeur = int(nouvelle_valeur)
            else:
                nouvelle_valeur = float(nouvelle_valeur)
        except:
            erreur = "⛔ Format numérique attendu"

    if not erreur:
        if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
            try:
                df.at[index, colonne] = nouvelle_valeur
            except Exception as e:
                erreur = f"⛔ {e}"
            else:
                df.at[index, colonne] = valeur_courante
                undo_redo_save()
                df.at[index, colonne] = nouvelle_valeur
                if inhiber_gestion_modification_cellule:
                    st.session_state.aggrid_activites_programmees_gerer_modification_cellule = False
                    st.session_state.aggrid_activites_non_programmees_gerer_modification_cellule = False
            
    return erreur


# Vérifie qu'une valeur est bien Oui Non
def est_reserve_valide(val):
    return str(val).strip().lower() in ["oui", "non", ""]

# Vérifie qu'une valeur contient bien NaN ou "" ou quelque chose du type "1", "1,10", "1, 10", "1, pair", "12, impair"
def est_relache_valide(val):

    # Cas val vide ou NaN
    if pd.isna(val) or str(val).strip() == "":
        return True

    val_str = str(val).strip().lower()

    # Autorise : chiffres ou mots-clés (pair, impair) séparés par virgules
    # Exemples valides : "1", "1, 10", "1, impair", "2, pair"
    # Regex : liste d'éléments séparés par des virgules, chaque élément est un entier ou 'pair'/'impair'
    motif = r"^\s*(\d+|pair|impair)(\s*,\s*(\d+|pair|impair))*\s*$"

    return re.fullmatch(motif, val_str) is not None

# Vérifie si une date de référence est compatible avec la valeur de la colonne Relache qui donne les jours de relache pour un spectacle donné
def est_hors_relache(relache_val, date_val):
    if pd.isna(relache_val) or pd.isna(date_val):
        return True  # Aucune relâche spécifiée ou date absente

    if not est_relache_valide(relache_val):
        return True
    
    try:
        date_int = int(float(date_val))
    except (ValueError, TypeError):
        return True  # Si la date n'est pas exploitable, on la considère programmable

    # Normaliser le champ Relache en chaîne
    if isinstance(relache_val, (int, float)):
        relache_str = str(int(relache_val))
    else:
        relache_str = str(relache_val).strip().lower()

    # Cas particulier : pair / impair
    if "pair" in relache_str and date_int % 2 == 0:
        return False
    if "impair" in relache_str and date_int % 2 != 0:
        return False

    # Cas général : liste explicite de jours (ex : "20,21")
    try:
        jours = [int(float(x.strip())) for x in relache_str.split(",")]
        if date_int in jours:
            return False
    except ValueError:
        pass  # ignorer s'il ne s'agit pas d'une liste de jours

    return True

# Suppression d'une activité d'un df
def supprimer_activite(df, idx):
    df.loc[idx] = pd.NA

# Déprogrammation d'une activité programmée d'un df (si pause suppression, si activité ordinaire date à None)
def deprogrammer_activite_programmee(df, idx):
    if est_pause(df.loc[idx]):
        df.loc[idx] = pd.NA
    else:
        df.at[idx, "Date"] = None

# Création de la liste des créneaux avant/après pour chaque activité programmée 
# le df des activités programmées est supposé etre trié par jour ("Date") et par heure de début ("Debut")
def get_creneaux(df, activites_programmees, traiter_pauses):

    def description_creneau(row, borne_min, borne_max, avant, apres, type_creneau):
        titre = row["Activite"] if not pd.isna(row["Activite"]) else ""
        date_str = str(int(row["Date"])) if pd.notnull(row["Date"]) else ""
        return {
            "Date": date_str,
            "Debut": borne_min.strftime('%Hh%M'),
            "Fin": borne_max.strftime('%Hh%M'),
            "Activité avant": avant,
            "Activité après": apres,
            "__type_creneau": type_creneau,
            "__index": row.name
        }
    
    params_to_hash = [
        traiter_pauses, 
        st.session_state.get("MARGE", MARGE).total_seconds(), 
        st.session_state.get("DUREE_REPAS", DUREE_REPAS).total_seconds(), 
        st.session_state.get("DUREE_CAFE", DUREE_CAFE).total_seconds()
    ]

    hash_val  = hash_df(df, colonnes_a_garder=[col for col in df.columns if col not in ["Debut_dt", "Duree_dt"]], params=params_to_hash)
    hash_key = "creneaux__hash"
    key = "creneaux"
    
    if st.session_state.get(hash_key) != hash_val:
        
        creneaux = []
        bornes = []
        jour_courant = activites_programmees.iloc[0]["Date"]

        for _, row in activites_programmees.iterrows():

            # Heure de début d'activité
            heure_debut = row["Debut_dt"]
            # Heure de fin d'activité
            heure_fin = heure_debut + row["Duree_dt"] if pd.notnull(heure_debut) and pd.notnull(row["Duree_dt"]) else None
            # initialisation du tableau enregistrant pour chaque jour les bornes des creneaux rencontrés pour eviter les doublons
            if row ["Date"] != jour_courant:
                bornes = []
                jour_courant = row ["Date"]

            # Ajout des creneaux avant l'activité considérée s'ils existent
            if pd.notnull(heure_debut):
                if get_activites_programmables_avant(df, activites_programmees, row, traiter_pauses):
                    borne_min, borne_max, pred = get_creneau_bounds_avant(activites_programmees, row)
                    if (borne_min, borne_max) not in bornes:
                        bornes.append((borne_min, borne_max))
                        creneaux.append(description_creneau(row, borne_min, borne_max, pred["Activite"] if pred is not None else "", row["Activite"], "Avant"))

            # Ajout des creneaux après l'activité considérée s'ils existent
            if pd.notnull(heure_fin):
                if get_activites_programmables_apres(df, activites_programmees, row, traiter_pauses):
                    borne_min, borne_max, next = get_creneau_bounds_apres(activites_programmees, row)
                    if (borne_min, borne_max) not in bornes:
                        bornes.append((borne_min, borne_max))
                        creneaux.append(description_creneau(row, borne_min, borne_max, row["Activite"], next["Activite"] if next is not None else "", "Après"))

        st.session_state[key] = pd.DataFrame(creneaux).sort_values(by=["Date", "Debut"], ascending=[True, True]) if creneaux else pd.DataFrame(creneaux)
        st.session_state[hash_key] = hash_val
    return st.session_state[key]

# Renvoie les bornes du créneau existant avant une activité donnée par son descripteur ligne_ref
def get_creneau_bounds_avant(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None    

    # Chercher l'activité programmée précédente sur le même jour
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    prev = programmes_jour_ref[programmes_jour_ref["Debut_dt"] < debut_ref].tail(1)

    # Calculer l'heure de début minimum du créneau
    if not prev.empty:
        prev_fin = datetime.datetime.combine(BASE_DATE, prev["Debut_dt"].iloc[0].time()) + prev["Duree_dt"].iloc[0]
        debut_min = prev_fin
    else:
        debut_min = datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))

    # Calculer l'heure de fin max du créneau
    fin_max = datetime.datetime.combine(BASE_DATE, debut_ref.time())

    return debut_min, fin_max, prev.iloc[0] if not prev.empty else None

# Renvoie les bornes du créneau existant après une activité donnée par son descripteur ligne_ref
def get_creneau_bounds_apres(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else debut_ref    


    # Ajuster la date de référence si le jour a changé
    if fin_ref.day != debut_ref.day:
        date_ref = date_ref + fin_ref.day - debut_ref.day  

    # Chercher l'activité programmée suivante sur le même jour de référence
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    next = programmes_jour_ref[programmes_jour_ref["Debut_dt"] + programmes_jour_ref["Duree_dt"] > fin_ref].head(1)

    # Calculer l'heure de fin max du créneau
    if not next.empty:
        fin_max = datetime.datetime.combine(BASE_DATE, next["Debut_dt"].iloc[0].time())
    else:
        fin_max = datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))

    # Calculer l'heure de début minimum du créneau
    debut_min = datetime.datetime.combine(BASE_DATE, fin_ref.time())

    return debut_min, fin_max, next.iloc[0] if not next.empty else None

# Renvoie la liste des activités programmables avant une activité donnée par son descripteur ligne_ref
def get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None

    proposables = [] 

    debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    if debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # Le spectacle doit commencer après debut_min et finir avant fin_max
        if h_debut >= debut_min + st.session_state.MARGE and h_fin <= fin_max - st.session_state.MARGE and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Date", "Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Avant")
    return proposables

# Renvoie la liste des activités programmables après une activité donnée par son descripteur ligne_ref
def get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None   

    proposables = []

    debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref)
    if debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    if fin_ref.day != debut_ref.day:
        return proposables  # Pas d'activités programmables après si le jour a changé

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # Le spectacle doit commencer après debut_min et finir avant fin_max
        if h_debut >= debut_min + st.session_state.MARGE and h_fin <= fin_max - st.session_state.MARGE and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Date", "Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Après")
    return proposables
    
# Vérifie si une pause d'un type donné est déjà présente pour un jour donné dans le dataframe des activités planiées
def pause_deja_existante(activites_programmees, jour, type_pause):
    activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour]
    return activites_programmes_du_jour["Activite"].astype(str).str.contains(type_pause, case=False, na=False).any() 

# Ajoute les pauses possibles (déjeuner, dîner, café) à une liste d'activités programmables pour une activité donnée par son descripteur ligne_ref
def ajouter_pauses(proposables, activites_programmees, ligne_ref, type_creneau):

    # Pause repas
    def ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, pause_debut_min, pause_debut_max, type_repas):
        if not pause_deja_existante(activites_programmees, date_ref, type_repas):
            if type_creneau == "Avant":
                h_dej = min(max(fin_max - st.session_state.DUREE_REPAS - st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and h_dej + st.session_state.MARGE <= fin_max:
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas
                    })
                    proposables.append(nouvelle_ligne)
            elif type_creneau == "Après":
                h_dej = min(max(debut_min + st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and h_dej + st.session_state.MARGE <= fin_max:
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas
                    })
                    proposables.append(nouvelle_ligne)
    
    def ajouter_pause_cafe(proposables, debut_min, fin_max):
        if not est_pause(ligne_ref):
            Lieu_ref = ligne_ref["Lieu"]
            if type_creneau == "Avant":
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_prev = activites_programmees.iloc[i - 1]["Lieu"] if i > 0 else None
                h_cafe = fin_max - st.session_state.DUREE_CAFE
                if not pd.isna(Lieu_ref_prev) and Lieu_ref == Lieu_ref_prev: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec le spectacle précédent 
                    if h_cafe >= debut_min: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": f"Pause café",
                            "__type_activite": "café"
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec le spectacle précédent sauf si debut_min = 0h00
                    marge_cafe = st.session_state.MARGE if debut_min != datetime.datetime.combine(BASE_DATE, datetime.time(0, 0)) else datetime.timedelta(minutes=0) 
                    if h_cafe >= debut_min + marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café"
                        })
                        proposables.append(nouvelle_ligne)
            elif type_creneau == "Après":
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_suiv = activites_programmees.iloc[i + 1]["Lieu"] if i < len(activites_programmees) - 1 else None
                h_cafe = debut_min
                if not pd.isna(Lieu_ref_suiv) and Lieu_ref == Lieu_ref_suiv: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec le spectacle suivant 
                    if h_cafe + st.session_state.DUREE_CAFE <= fin_max: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café"
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec le spectacle suivant sauf si fin_max = 23h59
                    marge_cafe = st.session_state.MARGE if fin_max != datetime.datetime.combine(BASE_DATE, datetime.time(23, 59)) else datetime.timedelta(minutes=0)
                    if h_cafe + st.session_state.DUREE_CAFE <= fin_max - marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café"
                        })
                        proposables.append(nouvelle_ligne)

    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None    

    def desc(h, duree, nom):
        # return f"{int(date_ref)} de {h.strftime('%Hh%M')} à {(h + duree).time().strftime('%Hh%M')} ({formatter_timedelta(duree)}) - {nom}"
        return f"{int(date_ref)} - {h.strftime('%Hh%M')} - {nom}"
    
    # Récupération des bornes du créneau
    if type_creneau == "Avant":
        debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    elif type_creneau == "Après":
        debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref)
    else:
        raise ValueError("type_creneau doit être 'Avant' ou 'Après'")

    # Pause déjeuner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DEJ_DEBUT_MIN, PAUSE_DEJ_DEBUT_MAX, "déjeuner")

    # Pause dîner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DIN_DEBUT_MIN, PAUSE_DIN_DEBUT_MAX, "dîner")

    # Pause café
    ajouter_pause_cafe(proposables, debut_min, fin_max)

def est_pause_str(val):
    valeurs = val.split()
    if not valeurs:
        return False
    return val.split()[0].lower() == "pause"

def est_pause(ligne_ref):
    val = str(ligne_ref["Activite"]).strip()
    return est_pause_str(val)

def est_pause_cafe(ligne_ref):
    if not est_pause(ligne_ref):
        return False
    val = str(ligne_ref["Activite"]).strip()
    valeurs = val.split()
    if not valeurs:
        return False
    if len(valeurs) < 2:
        return False
    return val.split()[0].lower() == "pause" and val.split()[1].lower() == "café"

def sauvegarder_contexte():

    def serialiser_contexte(df):
        # Réindexer proprement pour éviter les trous
        df_sorted = df.copy()
        df_sorted = df_sorted.sort_values(by=["Date", "Debut_dt"])
        df_sorted = df_sorted.reset_index(drop=True)
        df_sorted = df_sorted.drop(columns=["Debut_dt", "Duree_dt"], errors='ignore')

        # Récupération de la worksheet à traiter
        wb = st.session_state.get("wb")

        if wb is not None:
            ws = wb.worksheets[0]
            liens_activites = st.session_state.get("liens_activites")

            # Effacer le contenu de la feuille Excel existante
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None  # on garde le style, on efface juste la valeur
                    cell.hyperlink = None

            # Réinjecter les données du df dans la feuille Excel
            from copy import copy

            col_activite = None
            for cell in ws[1]:
                if cell.value and str(cell.value).strip().lower() in ["activité"]:
                    col_activite = cell.column
            source_font = ws.cell(row=1, column=1).font

            # Réécriture sans saut de ligne
            for i, (_, row) in enumerate(df_sorted.iterrows()):
                row_idx = i + 2  # ligne Excel (1-indexée + entête)
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    if pd.isna(value):
                        cell.value = None
                    else:
                        try:
                            # Conserve les entiers réels, sinon cast en string
                            v = int(value)
                            if str(v) == str(value).strip():
                                cell.value = v
                            else:
                                cell.value = value
                        except (ValueError, TypeError):
                            cell.value = value

                        # Ajout d'hyperlien pour la colonne Activite
                        if col_activite is not None:
                            if col_idx == col_activite and isinstance(liens_activites, dict):
                                lien = liens_activites.get(value)
                                if lien:
                                    cell.hyperlink = lien
                                    cell.font = Font(color="0000EE", underline="single")
                                else:
                                    cell.hyperlink = None
                                    cell.font = copy(source_font)   

            # Sauvegarde dans un buffer mémoire
            buffer = io.BytesIO()
            wb.save(buffer)
        else:
            # Sauvegarde dans un buffer mémoire
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df_sorted.to_excel(writer, index=False)

        # Revenir au début du buffer pour le téléchargement
        buffer.seek(0)
        return buffer

    # Version modale
    @st.dialog("Sauvegarder données")
    def show_dialog_sauvegarder_contexte(df, nom_fichier):
        st.markdown("Voulez-vous sauvegarder les données ?")
        col1, col2 = st.columns([1, 1])
        with col1:
            df_hash = hash_df(st.session_state.df, colonnes_a_enlever=["Debut_dt", "Duree_dt"])
            prev_hash = st.session_state.get("__contexte_hash")
            buffer = st.session_state.get("__contexte_buffer")

            if df_hash != prev_hash or buffer is None:
                # Le df a changé, on régénère le buffer
                buffer = serialiser_contexte(st.session_state.df)
                st.session_state["__contexte_hash"] = df_hash
                st.session_state["__contexte_buffer"] = buffer

            # Bouton de téléchargement
            if st.download_button(
                label="Valider",
                data=st.session_state["__contexte_buffer"],
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=CENTRER_BOUTONS
            ):
                st.rerun()
        with col2:
            if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
                st.rerun()

    # Version Non Modale
    if "df" in st.session_state:
        nom_fichier = st.session_state.fn if "fn" in st.session_state else "planning_avignon.xlsx"
        
        df_hash = hash_df(st.session_state.df, colonnes_a_enlever=["Debut_dt", "Duree_dt"])
        prev_hash = st.session_state.get("__contexte_hash")
        buffer = st.session_state.get("__contexte_buffer")

        if df_hash != prev_hash or buffer is None:
            # Le df a changé, on régénère le buffer
            buffer = serialiser_contexte(st.session_state.df)
            st.session_state["__contexte_hash"] = df_hash
            st.session_state["__contexte_buffer"] = buffer

        # Bouton de téléchargement
        st.download_button(
            label=LABEL_BOUTON_SAUVEGARDER,
            data=st.session_state["__contexte_buffer"],
            file_name=nom_fichier,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=CENTRER_BOUTONS
        )

# Programme une activité non programmée à une date donnée
def programmer_activite_non_programmee(df, date_ref, activite):

    type_activite = activite["__type_activite"]
    undo_redo_save()
    if type_activite == "ActiviteExistante":
        # Pour les spectacles, on programme la date et l'heure
        index = activite["__index"]
        df.at[index, "Date"] = date_ref
    elif type_activite == "déjeuner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)  # Ajouter à la fin du DataFrame
        df.at[index, "Date"] = date_ref
        df.at[index, "Debut"] = activite["Debut"]
        df.at[index, "Duree"] = formatter_timedelta(st.session_state.DUREE_REPAS)
        df.at[index, "Activite"] = "Pause déjeuner"
    elif type_activite == "dîner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)  # Ajouter à la fin du DataFrame
        df.at[index, "Date"] = date_ref
        df.at[index, "Debut"] = activite["Debut"]
        df.at[index, "Duree"] = formatter_timedelta(st.session_state.DUREE_REPAS)
        df.at[index, "Activite"] = "Pause dîner"
    elif type_activite == "café":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)  # Ajouter à la fin du DataFrame
        df.at[index, "Date"] = date_ref
        df.at[index, "Debut"] = activite["Debut"]
        df.at[index, "Duree"] = formatter_timedelta(st.session_state.DUREE_CAFE)
        df.at[index, "Activite"] = "Pause café"

    st.session_state.activites_programmees_selected_row = index
    forcer_reaffichage_activites_programmees()
    forcer_reaffichage_df("creneaux_disponibles")
    # st.session_state.activites_non_programmees_selected_row = ligne_voisine_index(st.session_state.activites_non_programmees_df_display, index)
    # forcer_reaffichage_activites_non_programmees()

    sauvegarder_row_ds_gsheet(df, index)
    st.rerun()

# Renvoie les jours possibles pour programmer une activité donnée par son idx
def get_jours_possibles(df, activites_programmees, idx_activite):
    jours_possibles = []

    # Retour si index non valide
    if idx_activite not in df.index:
        return jours_possibles

    # Récupérer la durée de l'activité à considérer
    ligne_a_considerer = df.loc[idx_activite]
    debut = ligne_a_considerer["Debut_dt"]
    fin = ligne_a_considerer["Debut_dt"] + ligne_a_considerer["Duree_dt"]

    if activites_programmees is not None:
        for jour in range(st.session_state.periode_a_programmer_debut.day, st.session_state.periode_a_programmer_fin.day + 1):
            
            if not est_hors_relache(ligne_a_considerer["Relache"], jour):
                continue

            activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour].sort_values("Debut_dt")

            if not activites_programmes_du_jour.empty:
                # Créneau entre minuit et première activité du jour
                premiere_activite_du_jour = activites_programmes_du_jour.iloc[0]
                borne_inf = datetime.datetime.combine(BASE_DATE, datetime.time.min)  # 00h00
                borne_sup = premiere_activite_du_jour["Debut_dt"]
                if debut > borne_inf + st.session_state.MARGE and fin < borne_sup - st.session_state.MARGE:
                    jours_possibles.append(jour)
                    continue  # on prend le premier créneau dispo du jour

                # Ensuite, créneaux entre chaque activité programmée
                for _, ligne in activites_programmes_du_jour.iterrows():
                    borne_inf, borne_sup, _ = get_creneau_bounds_apres(activites_programmes_du_jour, ligne)
                    if debut > borne_inf + st.session_state.MARGE and fin < borne_sup - st.session_state.MARGE:
                        jours_possibles.append(jour)
                        break  # jour validé, on passe au suivant
            else: # jour libre
                jours_possibles.append(jour)

    return jours_possibles

# idem get_jours_possibles avec en paramètre une row d'activité programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees de get_jours_possibles sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jour_courant = int(row["Date"]) if pd.notna(row["Date"]) and row["Date"] is not None else row["Date"]
    if jours != []:
        if not est_reserve(st.session_state.df.loc[row["__index"]]):
            jours = [jour_courant] + jours + [""] 
        else: 
            jours = []
    return [str(j) for j in jours] if isinstance(jours, list) else []

# idem get_jours_possibles avec en paramètre une row d'activité non programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees de get_jours_possibles sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_non_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jours = [""] + jours if jours != [] else jours
    return [str(j) for j in jours] if isinstance(jours, list) else []

# Calcule les options des dates pour les activiés programmées
def calculer_options_date_activites_programmees(df_display):
    hash_val  = hash_df(df_display, colonnes_a_garder=["Date", "Debut", "Duree"])
    hash_key = "options_date_activites_programmees__hash"
    key = "options_date_activites_programmees"
    if st.session_state.get(hash_key) != hash_val:
        st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_programmee(row), axis=1)
        st.session_state[hash_key] = hash_val
    return st.session_state[key]

# Calcule les options des dates pour les activiés non programmées
def calculer_options_date_activites_non_programmees(df_display):
    hash_val  = hash_df(df_display, colonnes_a_garder=["Date", "Debut", "Duree"])
    hash_key = "options_date_activites_non_programmees__hash"
    key = "options_date_activites_non_programmees"
    if st.session_state.get(hash_key) != hash_val:
        st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_non_programmee(row), axis=1)
        st.session_state[hash_key] = hash_val
    return st.session_state[key]

# Programme une activité choisie en fonction des jours possibles
def programmer_activite_par_choix_activite(df):
    st.markdown("##### Programmation d'une nouvelle activité")

    # Filtrer les activités non programmées
    activites_programmees = get_activites_programmees(df)
    activites_non_programmees = get_activites_non_programmees(df)

    # Liste d'options formatées
    options_activites = []
    for idx, row in activites_non_programmees.iterrows():
        if get_jours_possibles(df, activites_programmees, idx):
            label = f"[{row["Debut"]} - {row["Fin"]}] - {str(row["Activite"]).strip()}"
            options_activites.append((label, idx))

    # Afficher la selectbox des activités
    activite_selectionee = st.selectbox("Choix de l'activité à programmer :", options_activites, format_func=lambda x: x[0])
    if activite_selectionee:
        idx_choisi = activite_selectionee[1]

        # Déterminer les jours disponibles 
        jours_possibles = get_jours_possibles(df, activites_programmees, idx_choisi)
        jours_label = [f"{int(jour):02d}" for jour in jours_possibles]

        jour_selection = st.selectbox("Choix du jour :", jours_label)

        # Bouton pour confirmer
        if jour_selection:
            if st.button(LABEL_BOUTON_PROGRAMMER, key="AjouterAuPlanningParChoixActivite"):
                jour_choisi = int(jour_selection.split()[-1])

                # On peut maintenant modifier le df
                df.at[idx_choisi, "Date"] = jour_choisi
                st.rerun()

# Programme une activité en fonction des créneaux possibles
def afficher_creneaux_disponibles(df):

    st.session_state.creneaux_disponibles_choix_activite = None
    activites_programmees = get_activites_programmees(df)
    if not activites_programmees.empty:

        # Rien à faire sur df vide
        if len(df) <= 0:
            return
        
        # initialisation du flag traiter_pauses
        if "traiter_pauses" not in st.session_state:
            st.session_state.traiter_pauses = False
        traiter_pauses = st.session_state.traiter_pauses

        # Création des créneaux avant/après pour chaque spectacle programmé
        creneaux = get_creneaux(df, activites_programmees, traiter_pauses) 

        if not creneaux.empty:
            st.markdown("##### Créneaux disponibles")
            choix_creneau_pred = st.session_state["creneaux_disponibles_selected_row"] if "creneaux_disponibles_selected_row" in st.session_state else None
            choix_creneau = afficher_df("Créneaux disponibles", creneaux, hide=["__type_creneau", "__index"], key="creneaux_disponibles")
            if choix_creneau is not None:
                if choix_creneau_pred is not None and choix_creneau_pred.to_dict() != choix_creneau.to_dict():
                    forcer_reaffichage_df("activites_programmables_dans_creneau_selectionne")
                type_creneau = choix_creneau["__type_creneau"]
                idx = choix_creneau["__index"]

                ligne_ref = activites_programmees.loc[idx]
                date_ref = ligne_ref["Date"]

                # Choix d'une activité à programmer dans le creneau choisi
                if type_creneau == "Avant":
                    proposables = get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses)

                elif type_creneau == "Après":
                    proposables = get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses)

                if proposables:
                    proposables = pd.DataFrame(proposables).sort_values(by=["Debut"], ascending=[True]) if proposables else pd.DataFrame(proposables)
                    label = f"Activités programmables sur le créneau du {int(date_ref)} entre [{choix_creneau["Debut"]}-{choix_creneau["Fin"]}]"
                    activite = afficher_df(label, proposables, hide=["__type_activite", "__index"], key="activites_programmables_dans_creneau_selectionne")

                    st.session_state.menu_creneaux_disponibles = {
                        "df": df,
                        "date": date_ref,
                        "creneau": choix_creneau,
                        "activite": activite,
                    }

                    if not st.session_state.get("sidebar_menus"):
                        with st.expander("Contrôles"):
                            menu_creneaux_disponibles(df, date_ref, choix_creneau, activite)
                    
                    return

    st.session_state.menu_creneaux_disponibles = {
        "df": df,
        "date": None,
        "creneau": None,
        "activite": None,
    }

# Menu de gestion des créneaux disponibles
def menu_creneaux_disponibles(df, date, creneau, activite):

    # Gestion du flag de traitement des pauses
    if "traiter_pauses" not in st.session_state: 
        st.session_state.traiter_pauses = False
    traiter_pauses = st.checkbox("Tenir compte des pauses", value=False)  
    if traiter_pauses != st.session_state.traiter_pauses:
        st.session_state.traiter_pauses = traiter_pauses
        forcer_reaffichage_df("creneaux_disponibles")
        st.session_state.creneaux_disponibles_choix_activite = None
        st.rerun()

    # Affichage du créneau sélectionné et de l'activité séléctionnées dans le créneau
    date = int(date) if est_float_valide(date) else ""

    if st.session_state.sidebar_menus:
        if creneau is not None and not creneau.empty:
            debut_creneau = creneau['Debut'] 
            activite_avant = creneau['Activité avant']
            fin_creneau = creneau['Fin'] 
            activite_apres = creneau['Activité après'] 
        else:
            debut_creneau = ""
            activite_avant = ""
            fin_creneau = ""
            activite_apres = ""

        if activite is not None and not activite.empty:
            debut_activite = activite['Debut']
            nom_activite = activite['Activite']
        else:
            debut_activite = ""
            nom_activite = ""

        st_info_error_avec_label("Le", f"{date}")
        st_info_error_avec_label(f"Entre {debut_creneau}", activite_avant)
        st_info_error_avec_label(f"Et {fin_creneau}", activite_apres)
        st_info_error_avec_label(f"A {debut_activite}", nom_activite)

    # Gestion du bouton Programmer
    if st.button(LABEL_BOUTON_PROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=activite is None, key="AjouterAuPlanningParCréneau"):
        programmer_activite_non_programmee(df, date, activite)

# Force le reaffichage de l'agrid des activités programmées
def forcer_reaffichage_activites_programmees():
    if "aggrid_activites_programmees_reset_counter" in st.session_state:
        st.session_state.aggrid_activites_programmees_reset_counter +=1 
    if "aggrid_activites_programmees_forcer_reaffichage" in st.session_state:
        st.session_state.aggrid_activites_programmees_forcer_reaffichage = True

# Force le reaffichage de l'agrid des activités non programmées
def forcer_reaffichage_activites_non_programmees():
    if "aggrid_activites_non_programmees_reset_counter" in st.session_state:
        st.session_state.aggrid_activites_non_programmees_reset_counter += 1 
    if "aggrid_activites_non_programmees_forcer_reaffichage" in st.session_state:
        st.session_state.aggrid_activites_non_programmees_forcer_reaffichage = True

# Initialisation des variables d'état du contexte après chargement des données du contexte
def initialiser_etat_contexte(df, wb, fn, lnk, ca):
    st.session_state.df = df
    st.session_state.wb = wb
    st.session_state.fn = fn
    st.session_state.liens_activites = lnk
    st.session_state.carnet_adresses = ca
    st.session_state.nouveau_fichier = True
    st.session_state.compteur_activite = 0
    st.session_state.menu_activites = None
    st.session_state.menu_activites_programmees = None
    st.session_state.menu_activites_non_programmees = None
    st.session_state.menu_creneaux_disponibles = None
    st.session_state.menu_activites = {
        "menu": "menu_activites_non_programmees",
        "df": df,
        "index_df": None,
        "df_display": None,
        "nom_activite": ""
    }
    st.session_state.forcage_menu_activites_enabled = False
    st.session_state.forcer_menu_activites_programmees = False
    st.session_state.forcer_menu_activites_non_programmees = False
    initialiser_periode_a_programmer(df)

    forcer_reaffichage_activites_programmees()
    forcer_reaffichage_activites_non_programmees()
    forcer_reaffichage_df("creneaux_disponibles")

# Ajout d'une nouvelle activité 
def afficher_bouton_nouvelle_activite(df, disabled=False, key="ajouter_activite"):
    import numpy as np

    def get_nom_nouvelle_activite(df):
        noms_existants = df["Activite"].dropna().astype(str).str.strip().tolist()
        while True:
            st.session_state.compteur_activite += 1
            nom_candidat = f"Activité {st.session_state.compteur_activite}"
            if nom_candidat not in noms_existants:
                return nom_candidat
            
    def get_next_free_index(df):
        existing = set(df.index)
        i = 0
        while i in existing:
            i += 1
        return i
    
    # Initialiser le DataFrame dans session_state si absent
    if "compteur_activite" not in st.session_state:
        st.session_state.compteur_activite = 0

    # Bouton Ajouter
    if st.button(LABEL_BOUTON_AJOUTER, use_container_width=CENTRER_BOUTONS, disabled=disabled, key=key):

        undo_redo_save()
        new_idx = get_next_free_index(df)
        df.at[new_idx, "Debut"] = "09h00"
        df.at[new_idx, "Duree"] = "1h00"
        nom_activite = get_nom_nouvelle_activite(df)
        df.at[new_idx, "Activite"] = nom_activite
        st.session_state.activites_non_programmees_selected_row = new_idx
        st.session_state.editeur_activite_courante_idx = new_idx
        
        if MENU_ACTIVITE_UNIQUE:
            # st.session_state.forcer_menu_activites_non_programmees = True
            # Bascule du menu activité sur le menu_activites_non_programmees
            st.session_state.menu_activites = {
                "menu": "menu_activites_non_programmees",
                "df": df,
                "index_df": new_idx,
                "df_display": st.session_state.activites_non_programmees_df_display,
                "nom_activite": nom_activite
            }

        forcer_reaffichage_activites_non_programmees()
        forcer_reaffichage_df("activites_programmables_dans_creneau_selectionne")
        sauvegarder_row_ds_gsheet(df, new_idx)
        st.rerun()

# Charge le fichier Excel contenant les activités à programmer
def charger_contexte_depuis_fichier():
    # Callback de st.file_uploader pour charger le fichier Excel
    def file_uploader_callback():
        st.session_state.contexte_invalide = True
        fd = st.session_state.get("file_uploader")
        if fd is not None:
            try:
                df = pd.read_excel(fd)
                wb = load_workbook(fd)
                lnk = get_liens_activites(wb)
                sheetnames = wb.sheetnames
                ca = pd.read_excel(fd, sheet_name=sheetnames[1]) if len(sheetnames) > 1 else None
                df = nettoyer_donnees(df)
                if "contexte_invalide" not in st.session_state:
                    initialiser_etat_contexte(df, wb, fd.name, lnk, ca)
                    undo_redo_init(verify=False)
                    sauvegarder_contexte_ds_gsheet(df, lnk, fd, ca)
            except Exception as e:
                st.sidebar.error(f"Erreur de chargement du fichier : {e}")
                st.session_state.contexte_invalide = True

    # Chargement du fichier Excel contenant les activités à programmer
    st.file_uploader(
        "Choix du fichier Excel contenant les activités à programmer", 
        type=["xlsx"], 
        label_visibility="collapsed",
        key="file_uploader",
        on_change=file_uploader_callback)

# Initialisation des types d'un df vide
def initialiser_dtypes(df):
    for col in df.columns:
        if col in ["Date", "Priorite"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        else:
            df[col] = df[col].astype("str")

# Initialisation d'un nouveau contexte
def initialiser_nouveau_contexte(avec_sauvegarde=True):
    df = pd.DataFrame(columns=COLONNES_ATTENDUES)
    initialiser_dtypes(df)
    wb = None
    fn = "planning_avignon.xlsx"
    lnk = {}
    ca = pd.DataFrame(columns=COLONNES_ATTENDUES_CARNET_ADRESSES)
    
    df["Date"] = pd.to_numeric(df["Date"], errors="coerce").astype("Int64")
    df["Priorite"] = pd.to_numeric(df["Priorite"], errors="coerce").astype("Int64")

    initialiser_etat_contexte(df, wb, fn, lnk, ca)
    if avec_sauvegarde:
        sauvegarder_contexte_ds_gsheet(df, lnk, ca)
    maj_donnees_calculees(df)

# Création d'un nouveau contexte
def creer_nouveau_contexte():
    if st.button(LABEL_BOUTON_NOUVEAU, use_container_width=CENTRER_BOUTONS, key="creer_nouveau_contexte"):
        undo_redo_save()
        initialiser_nouveau_contexte()
        st.rerun()

# Indique si le contexte est vlide pour traitement
def est_contexte_valide():
    return "df" in st.session_state and isinstance(st.session_state.df, pd.DataFrame) and "contexte_invalide" not in st.session_state

# Affichage des contrôles d'édition
def afficher_controles_edition():
    if st.button(LABEL_BOUTON_DEFAIRE, 
        disabled=not st.session_state.get("historique_undo"), 
        use_container_width=CENTRER_BOUTONS, 
        key="undo_btn") and st.session_state.historique_undo:
        undo_redo_undo()
    if st.button(LABEL_BOUTON_REFAIRE, 
        disabled=not st.session_state.get("historique_redo"), 
        use_container_width=CENTRER_BOUTONS, 
        key="redo_btn") and st.session_state.historique_redo:
        undo_redo_redo()

# Affichage des choix généraux
def afficher_infos_generales(df):
    with st.expander("ℹ️ Infos"):
        # Vérification de l'
        afficher_aide()        
        
        # Vérification de cohérence des informations du df
        verifier_coherence(df) 

        # Choix de la période à programmer
        afficher_periode_a_programmer(df)

        # Affichage des paramètres
        afficher_parametres()

# Initialisation de la page HTML
def initialiser_page():

    # Evite la surbrillance rose pâle des lignes qui ont le focus sans être sélectionnées dans les AgGrid
    patch_aggrid_css()

# Affiche le nom d'activité
def afficher_nom_activite(df, index_df, nom_activite=None, afficher_label=True):

    afficher_label = False if not st.session_state.sidebar_menus else afficher_label
    
    if index_df is not None:
        row = df.loc[index_df]
        if nom_activite == None:
            nom_activite = row["Activite"].strip()
        if est_activite_programmee(row):
            label_activite = f"Le {int(row["Date"])} de {row["Debut"]} à {row["Fin"]}"
            if est_reserve(row):
                st_info_error_avec_label(label_activite, nom_activite, afficher_label=afficher_label, color="red")
            else:
                st_info_error_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
        else:
            label_activite = f"De {row["Debut"]} à {row["Fin"]}"
            st_info_error_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
    else:
        if nom_activite == None:
            nom_activite = ""
        label_activite = "De ..h.. à ..h.."
        st_info_error_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
    
# Affiche un nom d'activité clickable qui switche le menu d'activités alternatif (sert en mode MODE_ACTIVITE_UNIQUE)
def afficher_nom_activite_clickable(df, index_df, nom_activite=None, afficher_label=True):

    hit = False
    key = "nom_activite_clickable" if st.session_state.sidebar_menus else None
    afficher_label = False if not st.session_state.sidebar_menus else afficher_label

    if index_df is not None:
        row = df.loc[index_df]
        activite_reserve = est_reserve(row)

        # Injecte le CSS permettent de styler le primary button affiché par st_info_error_avec_label avec param key 
        injecter_css_pour_primary_buttons("error" if activite_reserve else "info")

        if nom_activite == None:
            nom_activite = row["Activite"].strip()
        if est_activite_programmee(row):
            label_activite = f"Le {int(row["Date"])} de {row["Debut"]} à {row["Fin"]}"
            if activite_reserve:
                hit = st_info_error_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label, color="red")
            else:
                hit = st_info_error_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
        else:
            label_activite = f"De {row["Debut"]} à {row["Fin"]}"
            hit = st_info_error_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
    else:
        if nom_activite == None:
            nom_activite = ""
        label_activite = "De ..h.. à ..h.."

        # Injecte le CSS permettent de styler le primary button affiché par st_info_error_avec_label avec param key 
        injecter_css_pour_primary_buttons("info")
        hit = st_info_error_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
    
    if hit:
        if est_activite_programmee(df.loc[index_df]):
            new_index_df = st.session_state.activites_non_programmees_selected_row
            if new_index_df is not None:
                new_df_display = st.session_state.activites_non_programmees_df_display
                new_nom_activite = df.loc[new_index_df, "Activite"] 
                st.session_state.menu_activites = {
                    "menu": "menu_activites_non_programmees",
                    "df": df,
                    "index_df": new_index_df,
                    "df_display": new_df_display,
                    "nom_activite": new_nom_activite
                }
        else:
            new_index_df = st.session_state.activites_programmees_selected_row
            if new_index_df is not None:
                new_df_display = st.session_state.activites_programmees_df_display
                new_nom_activite = df.loc[new_index_df, "Activite"] 
                st.session_state.menu_activites = {
                    "menu": "menu_activites_programmees",
                    "df": df,
                    "index_df": new_index_df,
                    "df_display": new_df_display,
                    "nom_activite": new_nom_activite
                }
        st.rerun()

# Affichage de la la sidebar min avec menus fichier et edition 
# (le reste est affiché dans d'affichage de données en fonction du contexte)
def afficher_sidebar(df):
    st.sidebar.title("Menu principal")

    with st.sidebar.expander("Fichier"):
        charger_contexte_depuis_fichier()
        creer_nouveau_contexte()
        sauvegarder_contexte()

    with st.sidebar.expander("Edition"):
        afficher_controles_edition()

    if not st.session_state.get("sidebar_menus"):
        return
    
    if MENU_ACTIVITE_UNIQUE:
        with st.sidebar.expander("Activités"):
            if "menu_activites" in st.session_state and isinstance(st.session_state.menu_activites, dict):
                if st.session_state.menu_activites["menu"] == "menu_activites_programmees":
                    menu_activites_programmees(
                        st.session_state.menu_activites["df"],
                        st.session_state.menu_activites["index_df"],
                        st.session_state.menu_activites["df_display"],
                        st.session_state.menu_activites["nom_activite"],
                    )

                elif st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
                    menu_activites_non_programmees(
                        st.session_state.menu_activites["df"],
                        st.session_state.menu_activites["index_df"],
                        st.session_state.menu_activites["df_display"],
                        st.session_state.menu_activites["nom_activite"],
                    )

    else:
        with st.sidebar.expander("Activités programmées"):
            if "menu_activites_programmees" in st.session_state and isinstance(st.session_state.menu_activites_programmees, dict):
                menu_activites_programmees(
                    st.session_state.menu_activites_programmees["df"],
                    st.session_state.menu_activites_programmees["index_df"],
                    st.session_state.menu_activites_programmees["df_display"],
                    st.session_state.menu_activites_programmees["nom_activite"],
                )

        with st.sidebar.expander("Activités non programmées"):
            if "menu_activites_non_programmees" in st.session_state and isinstance(st.session_state.menu_activites_non_programmees, dict):
                menu_activites_non_programmees(
                    st.session_state.menu_activites_non_programmees["df"],
                    st.session_state.menu_activites_non_programmees["index_df"],
                    st.session_state.menu_activites_non_programmees["df_display"],
                    st.session_state.menu_activites_non_programmees["nom_activite"],
                )

    with st.sidebar.expander("Créneaux disponibles"):
        if "menu_creneaux_disponibles" in st.session_state and isinstance(st.session_state.menu_creneaux_disponibles, dict):
            menu_creneaux_disponibles(
                st.session_state.menu_creneaux_disponibles["df"],
                st.session_state.menu_creneaux_disponibles["date"],
                st.session_state.menu_creneaux_disponibles["creneau"],
                st.session_state.menu_creneaux_disponibles["activite"],
            )

def main():
    
    if "main_counter" not in st.session_state:
        st.session_state.main_counter = 0
    st.session_state.main_counter += 1
    debug_trace(f"____________MAIN {st.session_state.main_counter}______________", trace_type=["gen"])
    
    # Gestion du chargement de contexte depuis la Google Sheet en charge de la persistence 
    debug_trace("charger_contexte_depuis_gsheet", trace_type=["gen"])
    charger_contexte_depuis_gsheet()

    # Configuration de la page HTML
    debug_trace("initialiser_page", trace_type=["gen"])
    initialiser_page()

    # Affichage du titre
    debug_trace("afficher_titre", trace_type=["gen"])
    afficher_titre("Planificateur Avignon Off")

    # Si le contexte est valide, on le traite
    if est_contexte_valide():

        # Accès au DataFrame 
        df = st.session_state.df

        # Met à jour les données calculées
        debug_trace("maj_donnees_calculees", trace_type=["gen"])
        maj_donnees_calculees(df)

        # Affichage des infos générales
        debug_trace("afficher_infos_generales", trace_type=["gen"])
        afficher_infos_generales(df)
        
        # Affichage des activités programmées
        debug_trace("afficher_activites_programmees", trace_type=["gen"])
        afficher_activites_programmees(df)

        # Affichage des activités non programmées
        debug_trace("afficher_activites_non_programmees", trace_type=["gen"])
        afficher_activites_non_programmees(df)

        # Programmation d'une nouvelle activité par créneau
        debug_trace("afficher_creneaux_disponibles", trace_type=["gen"])
        afficher_creneaux_disponibles(df)            

        # Affichage de la sidebar
        debug_trace("afficher_sidebar", trace_type=["gen"])
        afficher_sidebar(df)

if __name__ == "__main__":
    main()
